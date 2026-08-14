#!/usr/bin/env python3
"""Recalculate the primary CheckMate analyses from Braun et al. source tables.

The script deliberately keeps the source OS_CNSR/PFS_CNSR values as the event
indicators used in the published table. In this dataset, value 1 behaves as an
observed event (short OS/PFS in progressing/deceased patients) and value 0 as
censoring. All inferential analyses are restricted to the 181 nivolumab-treated
RNA-profiled tumors and use the frozen C0-C9 signature definition in resources/.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
os.environ.setdefault("MPLCONFIGDIR", str(ROOT / ".cache" / "matplotlib"))

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from scipy import stats
from lifelines import CoxPHFitter, KaplanMeierFitter
from lifelines.statistics import logrank_test, proportional_hazard_test


SOURCE_CLINICAL_XLSX: Path
SOURCE_EXPRESSION_XLSX: Path
SIGNATURE_CSV = ROOT / "resources" / "CAR_T_state_signatures.csv"
OUT: Path

TCELL_GENES = ["CD2", "CD3D", "CD3E", "CD8A", "CD8B"]
COLORS = {
    "Tcell": "#D62728",
    "C0": "#4CAF50",
    "C1": "#F28E2B",
    "C2": "#8C564B",
    "C3": "#9467BD",
    "C4": "#17BECF",
    "C5": "#1F77B4",
    "C6": "#D62728",
    "C7": "#2A96E6",
    "C8": "#A27AE8",
    "C9": "#7F7F7F",
}


def verify_source_checksum(path: Path, source_id: str) -> None:
    manifest_path = ROOT / "data" / "source_manifest.tsv"
    with manifest_path.open("r", encoding="utf-8", newline="") as handle:
        records = {
            row["source_id"]: row
            for row in csv.DictReader(handle, delimiter="\t")
        }
    if source_id not in records:
        raise RuntimeError(f"Source manifest lacks {source_id!r}")
    record = records[source_id]
    expected_size = int(record["size_bytes"])
    observed_size = path.stat().st_size
    if observed_size != expected_size:
        raise RuntimeError(
            f"Size mismatch for {path}: expected {expected_size}, found {observed_size}"
        )
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    observed_hash = digest.hexdigest()
    if observed_hash != record["sha256"]:
        raise RuntimeError(
            f"SHA-256 mismatch for {path}: expected {record['sha256']}, "
            f"found {observed_hash}"
        )


def bh_adjust(values: pd.Series) -> pd.Series:
    arr = values.to_numpy(dtype=float)
    out = np.full(arr.shape, np.nan, dtype=float)
    ok = np.isfinite(arr)
    p = arr[ok]
    if p.size:
        order = np.argsort(p)
        ranked = p[order]
        adj = ranked * len(ranked) / np.arange(1, len(ranked) + 1)
        adj = np.minimum.accumulate(adj[::-1])[::-1]
        adj = np.minimum(adj, 1.0)
        restored = np.empty_like(adj)
        restored[order] = adj
        out[np.where(ok)[0]] = restored
    return pd.Series(out, index=values.index)


def read_source() -> tuple[pd.DataFrame, pd.DataFrame, dict[str, list[str]]]:
    if not SOURCE_CLINICAL_XLSX.exists():
        raise FileNotFoundError(SOURCE_CLINICAL_XLSX)
    if not SOURCE_EXPRESSION_XLSX.exists():
        raise FileNotFoundError(SOURCE_EXPRESSION_XLSX)
    if not SIGNATURE_CSV.exists():
        raise FileNotFoundError(SIGNATURE_CSV)

    clinical_wb = openpyxl.load_workbook(
        SOURCE_CLINICAL_XLSX, read_only=True, data_only=True
    )
    clinical_sheets = clinical_wb.sheetnames
    clinical_wb.close()
    clinical_sheet = next(
        (
            candidate
            for candidate in ["S1_Clinical_and_Immune_Data", "Clinical_and_Immune_Data"]
            if candidate in clinical_sheets
        ),
        None,
    )
    if clinical_sheet is None:
        raise RuntimeError(
            "The Braun clinical workbook lacks the expected S1 clinical sheet."
        )
    clinical = pd.read_excel(
        SOURCE_CLINICAL_XLSX, sheet_name=clinical_sheet, header=1
    )
    clinical = clinical.loc[clinical["RNA_ID"].notna()].copy()
    clinical["RNA_ID"] = clinical["RNA_ID"].astype(str)
    if clinical["RNA_ID"].duplicated().any():
        raise RuntimeError("Braun clinical table contains duplicate RNA_ID values.")
    if len(clinical) != 311:
        raise RuntimeError(f"Expected 311 RNA-profiled tumors; found {len(clinical)}.")
    clinical["OS_event"] = pd.to_numeric(clinical["OS_CNSR"], errors="coerce")
    clinical["PFS_event"] = pd.to_numeric(clinical["PFS_CNSR"], errors="coerce")
    for col in ["OS", "PFS", "OS_event", "PFS_event"]:
        clinical[col] = pd.to_numeric(clinical[col], errors="coerce")

    marker_df = pd.read_csv(SIGNATURE_CSV)
    required_marker_cols = {"cluster", "gene"}
    if not required_marker_cols.issubset(marker_df.columns):
        raise RuntimeError("Signature CSV lacks cluster/gene columns")
    marker_df["signature"] = "C" + marker_df["cluster"].astype(int).astype(str)
    marker_df = marker_df.loc[marker_df["signature"].isin(
        [f"C{i}" for i in range(10)]
    )].copy()
    marker_sets = {
        signature: group["gene"].dropna().astype(str).drop_duplicates().tolist()[:20]
        for signature, group in marker_df.groupby("signature", sort=False)
    }
    if set(marker_sets) != {f"C{i}" for i in range(10)}:
        raise RuntimeError("Signature CSV must define C0-C9")
    if any(len(genes) != 20 for genes in marker_sets.values()):
        raise RuntimeError("Each frozen C0-C9 signature must contain exactly 20 genes")

    wanted = sorted(
        set(TCELL_GENES)
        .union({"ICAM1", "B2M"})
        .union(*map(set, marker_sets.values()))
    )

    wb = openpyxl.load_workbook(
        SOURCE_EXPRESSION_XLSX, read_only=True, data_only=True
    )
    expression_sheet = next(
        (
            candidate
            for candidate in ["S4A_RNA_Expression", "RNA_Expression"]
            if candidate in wb.sheetnames
        ),
        None,
    )
    if expression_sheet is None:
        raise RuntimeError(
            "The Braun expression workbook lacks the expected S4 RNA sheet."
        )
    ws = wb[expression_sheet]
    header = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    sample_names = [str(x) if x is not None else "" for x in header[1:]]
    nonempty_sample_names = [name for name in sample_names if name]
    if len(nonempty_sample_names) != len(set(nonempty_sample_names)):
        raise RuntimeError("Braun expression matrix contains duplicate RNA_ID columns.")
    sample_to_index = {sample: i + 1 for i, sample in enumerate(sample_names)}
    missing_samples = sorted(set(clinical["RNA_ID"]) - set(sample_to_index))
    if missing_samples:
        raise RuntimeError(f"RNA samples absent from expression matrix: {missing_samples[:5]}")

    selected_samples = clinical["RNA_ID"].tolist()
    selected_indices = [sample_to_index[x] for x in selected_samples]
    gene_rows: dict[str, list[np.ndarray]] = {}
    wanted_set = set(wanted)
    for row in ws.iter_rows(min_row=3, values_only=True):
        gene = row[0]
        if gene not in wanted_set:
            continue
        vals = np.array(
            [pd.to_numeric(row[idx], errors="coerce") for idx in selected_indices],
            dtype=float,
        )
        gene_rows.setdefault(str(gene), []).append(vals)

    expression_columns = {
        gene: np.nanmean(np.vstack(rows), axis=0)
        for gene, rows in gene_rows.items()
    }
    expr = pd.DataFrame(expression_columns, index=selected_samples)
    expr.index.name = "RNA_ID"
    expr = expr.reset_index()
    wb.close()

    clinical.to_csv(OUT / "checkmate_clinical_source_qc.csv", index=False)
    expr.to_csv(OUT / "checkmate_selected_expression_qc.csv", index=False)
    pd.DataFrame(
        [
            {"signature": name, "gene": gene, "present_in_bulk": gene in expr.columns}
            for name, genes in marker_sets.items()
            for gene in genes
        ]
    ).to_csv(OUT / "signature_gene_coverage.csv", index=False)
    return clinical, expr, marker_sets


def safe_z(series: pd.Series) -> pd.Series:
    x = pd.to_numeric(series, errors="coerce")
    sd = x.std(ddof=1)
    if not np.isfinite(sd) or sd == 0:
        return pd.Series(np.zeros(len(x)), index=x.index)
    return (x - x.mean()) / sd


def build_scores(
    clinical: pd.DataFrame,
    expr: pd.DataFrame,
    marker_sets: dict[str, list[str]],
    arm: str | None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    meta = clinical.copy()
    if arm is not None:
        meta = meta.loc[meta["Arm"].eq(arm)].copy()
    df = meta.merge(expr, on="RNA_ID", how="inner", validate="one_to_one")

    score_sets = {"Tcell": TCELL_GENES, **marker_sets}
    coverage = []
    all_score_genes = sorted(set().union(*map(set, score_sets.values())))
    present = [g for g in all_score_genes if g in df.columns]
    df.loc[:, present] = df[present].apply(safe_z, axis=0)
    score_columns: dict[str, pd.Series] = {}
    for name, genes in score_sets.items():
        genes_use = [g for g in genes if g in df.columns]
        coverage.append(
            {
                "analysis_arm": arm or "ALL",
                "signature": name,
                "n_genes_defined": len(genes),
                "n_genes_present": len(genes_use),
                "genes_present": ";".join(genes_use),
            }
        )
        if len(genes_use) < 3:
            score_columns[f"{name}_score"] = pd.Series(np.nan, index=df.index)
        else:
            score_columns[f"{name}_score"] = df[genes_use].mean(axis=1, skipna=True)
    df = pd.concat([df, pd.DataFrame(score_columns, index=df.index)], axis=1)
    return df, pd.DataFrame(coverage)


def fit_one(
    df: pd.DataFrame,
    score: str,
    endpoint: str,
    stratify_trial: bool = True,
) -> tuple[dict, pd.DataFrame]:
    time_col = "OS" if endpoint == "OS" else "PFS"
    event_col = "OS_event" if endpoint == "OS" else "PFS_event"
    score_col = f"{score}_score"
    cols = [time_col, event_col, score_col, "Cohort", "RNA_ID"]
    dat = df[cols].dropna().copy()
    dat[event_col] = dat[event_col].astype(int)
    if not set(dat[event_col].unique()).issubset({0, 1}):
        raise ValueError(f"Unexpected event coding in {event_col}")
    median = dat[score_col].median()
    dat["group"] = (dat[score_col] > median).astype(int)
    dat["score_z"] = safe_z(dat[score_col])

    low = dat.loc[dat["group"].eq(0)]
    high = dat.loc[dat["group"].eq(1)]
    lr = logrank_test(
        low[time_col],
        high[time_col],
        event_observed_A=low[event_col],
        event_observed_B=high[event_col],
    )

    cox_cols = [time_col, event_col, "group"]
    strata = None
    if stratify_trial and dat["Cohort"].nunique() > 1:
        cox_cols.append("Cohort")
        strata = ["Cohort"]
    cph = CoxPHFitter()
    cph.fit(
        dat[cox_cols],
        duration_col=time_col,
        event_col=event_col,
        strata=strata,
        formula="group",
    )
    s = cph.summary.loc["group"]

    cont_cols = [time_col, event_col, "score_z"]
    if strata:
        cont_cols.append("Cohort")
    cph_cont = CoxPHFitter()
    cph_cont.fit(
        dat[cont_cols],
        duration_col=time_col,
        event_col=event_col,
        strata=strata,
        formula="score_z",
    )
    sc = cph_cont.summary.loc["score_z"]
    try:
        ph = proportional_hazard_test(cph, dat[cox_cols], time_transform="rank")
        ph_p = float(ph.summary.loc["group", "p"])
    except Exception:
        ph_p = np.nan

    result = {
        "endpoint": endpoint,
        "signature": score,
        "n": len(dat),
        "events": int(dat[event_col].sum()),
        "n_low": len(low),
        "n_high": len(high),
        "median_score": float(median),
        "logrank_p": float(lr.p_value),
        "HR_high_vs_low": float(s["exp(coef)"]),
        "CI_low": float(s["exp(coef) lower 95%"]),
        "CI_high": float(s["exp(coef) upper 95%"]),
        "cox_p": float(s["p"]),
        "continuous_HR_per_SD": float(sc["exp(coef)"]),
        "continuous_CI_low": float(sc["exp(coef) lower 95%"]),
        "continuous_CI_high": float(sc["exp(coef) upper 95%"]),
        "continuous_cox_p": float(sc["p"]),
        "PH_test_p": ph_p,
        "trial_stratified_cox": bool(strata),
    }
    return result, dat


def plot_km(dat: pd.DataFrame, result: dict, arm_label: str, out_path: Path) -> None:
    endpoint = result["endpoint"]
    time_col = "OS" if endpoint == "OS" else "PFS"
    event_col = "OS_event" if endpoint == "OS" else "PFS_event"
    score = result["signature"]
    fig, ax = plt.subplots(figsize=(5.2, 4.5))
    for group, label, color in [
        (0, "Low", "#BDBDBD"),
        (1, "High", COLORS.get(score, "#D62728")),
    ]:
        part = dat.loc[dat["group"].eq(group)]
        km = KaplanMeierFitter(label=f"{label} (n={len(part)})")
        km.fit(part[time_col], part[event_col])
        km.plot_survival_function(ax=ax, ci_show=False, color=color, linewidth=2.2)
    ax.set_xlim(0, 75)
    ax.set_ylim(0, 1.02)
    ax.set_xlabel(f"{endpoint} (months)")
    ax.set_ylabel("Survival probability")
    readable_arm = {"nivolumab_primary": "nivolumab-treated ccRCC"}.get(
        arm_label, arm_label
    )
    ax.set_title(f"{score} score — {readable_arm}", fontweight="bold")
    lines = [f"log-rank p = {result['logrank_p']:.3g}"]
    if score.startswith("C") and np.isfinite(result.get("logrank_BH_10_states", np.nan)):
        lines.append(f"BH p (10 states) = {result['logrank_BH_10_states']:.3g}")
    txt = "\n".join(lines) + "\n" + (
        f"trial-stratified HR = {result['HR_high_vs_low']:.2f}\n"
        f"95% CI {result['CI_low']:.2f}–{result['CI_high']:.2f}\n"
        f"events = {result['events']}/{result['n']}"
    )
    ax.text(0.04, 0.06, txt, transform=ax.transAxes, fontsize=9, va="bottom")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.legend(frameon=False, loc="upper right")
    fig.tight_layout()
    fig.savefig(out_path, dpi=600, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def analyze_arm(
    clinical: pd.DataFrame,
    expr: pd.DataFrame,
    marker_sets: dict[str, list[str]],
    arm: str | None,
    label: str,
) -> pd.DataFrame:
    scored, coverage = build_scores(clinical, expr, marker_sets, arm)
    coverage.to_csv(OUT / f"coverage_{label}.csv", index=False)
    scored.to_csv(OUT / f"scores_{label}.csv", index=False)
    results = []
    plot_data = {}
    for endpoint in ["OS", "PFS"]:
        for signature in ["Tcell"] + [f"C{i}" for i in range(10)]:
            result, dat = fit_one(scored, signature, endpoint)
            result["analysis"] = label
            result["arm"] = arm or "ALL"
            results.append(result)
            plot_data[(endpoint, signature)] = (result, dat)
    res = pd.DataFrame(results)
    for endpoint in ["OS", "PFS"]:
        mask = res["endpoint"].eq(endpoint) & res["signature"].str.match(r"C\d+")
        res.loc[mask, "logrank_BH_10_states"] = bh_adjust(res.loc[mask, "logrank_p"])
        res.loc[mask, "cox_BH_10_states"] = bh_adjust(res.loc[mask, "cox_p"])
        res.loc[mask, "continuous_cox_BH_10_states"] = bh_adjust(
            res.loc[mask, "continuous_cox_p"]
        )
    res.to_csv(OUT / f"survival_results_{label}.csv", index=False)
    for endpoint, signature in [
        ("OS", "Tcell"),
        ("OS", "C0"),
        ("OS", "C1"),
        ("OS", "C6"),
        ("PFS", "C0"),
        ("PFS", "C6"),
    ]:
        _, dat = plot_data[(endpoint, signature)]
        result = res.loc[
            res["endpoint"].eq(endpoint) & res["signature"].eq(signature)
        ].iloc[0].to_dict()
        plot_km(
            dat,
            result,
            label,
            OUT / f"KM_{label}_{endpoint}_{signature}.png",
        )
    return res


def ols_term(y: np.ndarray, x: np.ndarray, term_index: int) -> dict[str, float]:
    """Return coefficient statistics for one OLS design-matrix column."""
    keep = np.isfinite(y) & np.all(np.isfinite(x), axis=1)
    y = y[keep]
    x = x[keep]
    beta, _, _, _ = np.linalg.lstsq(x, y, rcond=None)
    residual = y - x @ beta
    df_resid = len(y) - x.shape[1]
    sigma2 = float(np.sum(residual**2) / df_resid)
    cov = sigma2 * np.linalg.pinv(x.T @ x)
    se = float(np.sqrt(cov[term_index, term_index]))
    value = float(beta[term_index])
    t_value = value / se
    p = float(2 * stats.t.sf(abs(t_value), df=df_resid))
    crit = float(stats.t.ppf(0.975, df=df_resid))
    return {
        "n": int(len(y)),
        "beta": value,
        "se": se,
        "ci_low": value - crit * se,
        "ci_high": value + crit * se,
        "p": p,
        "df_resid": int(df_resid),
    }


def analyze_figure5g_nivolumab(
    clinical: pd.DataFrame,
    expr: pd.DataFrame,
    marker_sets: dict[str, list[str]],
) -> pd.DataFrame:
    """Recalculate the B2M+ICAM1/signature panel in nivolumab samples."""
    meta = clinical.loc[clinical["Arm"].eq("NIVOLUMAB")].copy()
    df = meta.merge(expr, on="RNA_ID", how="inner", validate="one_to_one")
    excluded = set(TCELL_GENES + ["ICAM1", "B2M"])
    clean_sets = {
        name: [gene for gene in genes if gene not in excluded]
        for name, genes in marker_sets.items()
    }
    genes = sorted(
        set(TCELL_GENES + ["ICAM1", "B2M"]).union(*map(set, clean_sets.values()))
    )
    genes_present = [gene for gene in genes if gene in df.columns]
    df.loc[:, genes_present] = df[genes_present].apply(safe_z, axis=0)
    state_columns: dict[str, pd.Series] = {}
    for name, signature in clean_sets.items():
        use = [gene for gene in signature if gene in df.columns]
        state_columns[f"{name}_state"] = safe_z(df[use].mean(axis=1))
    state_columns["TcellScore_z"] = safe_z(
        df[[g for g in TCELL_GENES if g in df]].mean(axis=1)
    )
    state_columns["ICAM1_z"] = safe_z(df["ICAM1"])
    state_columns["B2M_z"] = safe_z(df["B2M"])
    state_columns["combined_B2M_ICAM1_z"] = safe_z(
        (state_columns["ICAM1_z"] + state_columns["B2M_z"]) / 2
    )
    df = pd.concat([df, pd.DataFrame(state_columns, index=df.index)], axis=1)

    trial_dummies = pd.get_dummies(df["Cohort"], prefix="trial", drop_first=True, dtype=float)
    results = []
    predictors = ["ICAM1_z", "B2M_z", "combined_B2M_ICAM1_z"]
    for predictor in predictors:
        for name in [f"C{i}" for i in range(10)]:
            design = pd.concat(
                [
                    pd.Series(1.0, index=df.index, name="intercept"),
                    df[predictor],
                    df["TcellScore_z"],
                    trial_dummies,
                ],
                axis=1,
            )
            fit = ols_term(
                df[f"{name}_state"].to_numpy(float),
                design.to_numpy(float),
                term_index=1,
            )
            fit.update(
                {
                    "cluster": name,
                    "predictor": predictor,
                    "covariates": "TcellScore_z + trial",
                    "n_signature_genes": len(
                        [gene for gene in clean_sets[name] if gene in df.columns]
                    ),
                }
            )
            results.append(fit)
    out = pd.DataFrame(results)
    out["BH_p_30_models"] = bh_adjust(out["p"])
    out.to_csv(OUT / "Figure_5G_nivolumab_associations.csv", index=False)

    plot_out = out.loc[out["predictor"].eq("combined_B2M_ICAM1_z")].copy()
    fig, ax = plt.subplots(figsize=(8.5, 5.5))
    x = np.arange(len(plot_out))
    colors = [COLORS[c] for c in plot_out["cluster"]]
    ax.bar(x, plot_out["beta"], color=colors, width=0.78)
    ax.errorbar(
        x,
        plot_out["beta"],
        yerr=np.vstack([
            plot_out["beta"] - plot_out["ci_low"],
            plot_out["ci_high"] - plot_out["beta"],
        ]),
        fmt="none",
        ecolor="black",
        elinewidth=0.8,
        capsize=3,
    )
    ax.axhline(0, color="black", linestyle="--", linewidth=0.8)
    ax.set_ylim(-0.08, 0.72)
    for idx, (_, row) in enumerate(plot_out.iterrows()):
        y = row["ci_high"] + 0.035 if row["beta"] >= 0 else row["ci_low"] - 0.035
        va = "bottom" if row["beta"] >= 0 else "top"
        p_label = (
            "<0.001" if row["BH_p_30_models"] < 0.001
            else f"={row['BH_p_30_models']:.3f}"
        )
        ax.text(idx, y, f"BH p{p_label}", ha="center", va=va, fontsize=7)
    ax.set_xticks(x, plot_out["cluster"], fontweight="bold")
    ax.set_ylabel("Adjusted standardized beta")
    ax.set_xlabel("Transferred in-vitro T-cell-state score")
    ax.set_title(
        "B2M+ICAM1 association with transferred T-cell-state scores\n"
        "nivolumab-treated ccRCC; adjusted for T-cell score and trial; BH across 30 models",
        fontweight="bold",
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    fig.savefig(
        OUT / "Figure_5G_nivolumab_B2M_ICAM1_associations.png",
        dpi=600,
        bbox_inches="tight",
        facecolor="white",
    )
    plt.close(fig)
    return out


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--braun-workbook", type=Path,
        help="Official combined Braun supplement containing S1 and S4A sheets.",
    )
    parser.add_argument(
        "--braun-clinical", type=Path,
        help="Official split Braun Table S1 workbook.",
    )
    parser.add_argument(
        "--braun-expression", type=Path,
        help="Official split Braun Table S4 workbook.",
    )
    parser.add_argument(
        "--output-dir", type=Path, default=ROOT / "results" / "checkmate_validation"
    )
    args = parser.parse_args()
    if args.braun_workbook is not None:
        if args.braun_clinical is not None or args.braun_expression is not None:
            parser.error(
                "Use either --braun-workbook or the two split-workbook arguments."
            )
        clinical_source = expression_source = args.braun_workbook
    else:
        if args.braun_clinical is None or args.braun_expression is None:
            parser.error(
                "Provide --braun-workbook, or both --braun-clinical and "
                "--braun-expression."
            )
        clinical_source = args.braun_clinical
        expression_source = args.braun_expression

    global SOURCE_CLINICAL_XLSX, SOURCE_EXPRESSION_XLSX, OUT
    SOURCE_CLINICAL_XLSX = clinical_source.resolve()
    SOURCE_EXPRESSION_XLSX = expression_source.resolve()
    if SOURCE_CLINICAL_XLSX == SOURCE_EXPRESSION_XLSX:
        verify_source_checksum(
            SOURCE_CLINICAL_XLSX, "braun_checkmate_supplement"
        )
    else:
        verify_source_checksum(
            SOURCE_CLINICAL_XLSX, "braun_checkmate_clinical_split"
        )
        verify_source_checksum(
            SOURCE_EXPRESSION_XLSX, "braun_checkmate_expression_split"
        )
    OUT = args.output_dir.resolve()
    OUT.mkdir(parents=True, exist_ok=True)
    clinical, expr, marker_sets = read_source()

    arm_counts = clinical["Arm"].value_counts().to_dict()
    if arm_counts != {"NIVOLUMAB": 181, "EVEROLIMUS": 130}:
        raise RuntimeError(f"Unexpected Braun treatment-arm counts: {arm_counts}")
    nivolumab = clinical.loc[clinical["Arm"].eq("NIVOLUMAB")]
    trial_counts = nivolumab["Cohort"].value_counts().to_dict()
    expected_trials = {"CM-025": 120, "CM-010": 45, "CM-009": 16}
    if trial_counts != expected_trials:
        raise RuntimeError(f"Unexpected nivolumab trial counts: {trial_counts}")
    if int(clinical["OS_event"].sum()) != 231 or int(clinical["PFS_event"].sum()) != 276:
        raise RuntimeError("Unexpected all-arm event counts in the Braun source table.")
    if int(nivolumab["OS_event"].sum()) != 123 or int(nivolumab["PFS_event"].sum()) != 159:
        raise RuntimeError("Unexpected nivolumab event counts in the Braun source table.")

    qc = {
        "rna_samples": int(len(clinical)),
        "arm_counts": arm_counts,
        "trial_by_arm": pd.crosstab(clinical["Cohort"], clinical["Arm"]).to_dict(),
        "OS_CNSR_counts": clinical["OS_CNSR"].value_counts(dropna=False).to_dict(),
        "PFS_CNSR_counts": clinical["PFS_CNSR"].value_counts(dropna=False).to_dict(),
        "event_coding_used": "source CNSR value 1 = observed event; 0 = censored",
    }
    with open(OUT / "source_qc.json", "w", encoding="utf-8") as handle:
        json.dump(qc, handle, ensure_ascii=False, indent=2, default=str)

    nivo = analyze_arm(clinical, expr, marker_sets, "NIVOLUMAB", "nivolumab_primary")
    analyze_figure5g_nivolumab(clinical, expr, marker_sets)
    key = nivo.loc[
        nivo["signature"].isin(["Tcell", "C0", "C1", "C6"])
        & nivo["endpoint"].isin(["OS", "PFS"])
    ].copy()
    key.to_csv(OUT / "survival_key_results.csv", index=False)
    print(key.to_string(index=False))


if __name__ == "__main__":
    main()
