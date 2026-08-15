#!/usr/bin/env python3
"""Convert the Macrogen workbook to canonical bulk RNA-seq count tables."""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import io
import json
import math
import statistics
from collections import Counter, defaultdict
from itertools import combinations
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook


EXPECTED_SOURCE_SHA256 = (
    "41cee68a4cd33f72268b46dc78fd0708dd62655814ae19a8a39fefd8aa5d5989"
)
SOURCE_FILENAME_AS_RECEIVED = "Expression_Profile.GRCh38.gene(20260814-181921).xlsx"
EXPECTED_SHEET = "Expression_Profile.GRCh38.gene"
ANNOTATION_COLUMNS = (
    "Gene_ID",
    "Transcript_ID",
    "Gene_Symbol",
    "Description",
    "gene_biotype",
    "Protein_ID",
    "HGNC",
    "MIM",
    "Ensembl",
    "IMGT/GENE-DB",
)

SAMPLE_SPECS = (
    ("IFN__1_Read_Count", "WT_IFNG_R1", "WT", "", "IFNG", 1),
    ("IFN__2_Read_Count", "WT_IFNG_R2", "WT", "", "IFNG", 2),
    ("IFN__3_Read_Count", "WT_IFNG_R3", "WT", "", "IFNG", 3),
    ("TNF__1_Read_Count", "WT_TNF_R1", "WT", "", "TNF", 1),
    ("TNF__2_Read_Count", "WT_TNF_R2", "WT", "", "TNF", 2),
    ("TNF__3_Read_Count", "WT_TNF_R3", "WT", "", "TNF", 3),
    ("TI__1_Read_Count", "WT_TNF_IFNG_R1", "WT", "", "TNF_IFNG", 1),
    ("TI__2_Read_Count", "WT_TNF_IFNG_R2", "WT", "", "TNF_IFNG", 2),
    ("TI__3_Read_Count", "WT_TNF_IFNG_R3", "WT", "", "TNF_IFNG", 3),
    ("HELA_1_Read_Count", "WT_control_R1", "WT", "", "control", 1),
    ("HELA_2_Read_Count", "WT_control_R2", "WT", "", "control", 2),
    ("HELA_3_Read_Count", "WT_control_R3", "WT", "", "control", 3),
    ("T6_IFN__1_Read_Count", "TNFR1_KO1_IFNG_R1", "TNFR1_KO1", "T6", "IFNG", 1),
    ("T6_IFN__2_Read_Count", "TNFR1_KO1_IFNG_R2", "TNFR1_KO1", "T6", "IFNG", 2),
    ("T6_IFN__3_Read_Count", "TNFR1_KO1_IFNG_R3", "TNFR1_KO1", "T6", "IFNG", 3),
    (
        "T6_TI__1_Read_Count",
        "TNFR1_KO1_TNF_IFNG_R1",
        "TNFR1_KO1",
        "T6",
        "TNF_IFNG",
        1,
    ),
    (
        "T6_TI__2_Read_Count",
        "TNFR1_KO1_TNF_IFNG_R2",
        "TNFR1_KO1",
        "T6",
        "TNF_IFNG",
        2,
    ),
    (
        "T6_TI__3_Read_Count",
        "TNFR1_KO1_TNF_IFNG_R3",
        "TNFR1_KO1",
        "T6",
        "TNF_IFNG",
        3,
    ),
    ("T6_TNF__1_Read_Count", "TNFR1_KO1_TNF_R1", "TNFR1_KO1", "T6", "TNF", 1),
    ("T6_TNF__2_Read_Count", "TNFR1_KO1_TNF_R2", "TNFR1_KO1", "T6", "TNF", 2),
    ("T6_TNF__3_Read_Count", "TNFR1_KO1_TNF_R3", "TNFR1_KO1", "T6", "TNF", 3),
    ("T6_1_Read_Count", "TNFR1_KO1_control_R1", "TNFR1_KO1", "T6", "control", 1),
    ("T6_2_Read_Count", "TNFR1_KO1_control_R2", "TNFR1_KO1", "T6", "control", 2),
    ("T6_3_Read_Count", "TNFR1_KO1_control_R3", "TNFR1_KO1", "T6", "control", 3),
)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_count(value: object, row_number: int, column_name: str) -> int:
    if isinstance(value, bool) or value is None:
        raise ValueError(f"Missing or invalid count at row {row_number}, {column_name}")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Non-numeric count at row {row_number}, {column_name}: {value!r}"
        ) from exc
    if not number.is_integer() or number < 0:
        raise ValueError(
            f"Counts must be non-negative integers; row {row_number}, "
            f"{column_name} contains {value!r}"
        )
    return int(number)


def write_tsv(path: Path, header: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow(header)
        writer.writerows(rows)


def write_tsv_gz(
    path: Path, header: Sequence[str], rows: Iterable[Sequence[object]]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("wb") as raw:
        with gzip.GzipFile(filename="", mode="wb", fileobj=raw, mtime=0) as zipped:
            with io.TextIOWrapper(zipped, encoding="utf-8", newline="") as text_handle:
                writer = csv.writer(text_handle, delimiter="\t", lineterminator="\n")
                writer.writerow(header)
                writer.writerows(rows)


def log1p_pearson(records: Sequence[tuple[int, list[str], list[int]]], a: int, b: int) -> float:
    n = len(records)
    sum_x = sum_y = sum_xx = sum_yy = sum_xy = 0.0
    for _, _, counts in records:
        x = math.log1p(counts[a])
        y = math.log1p(counts[b])
        sum_x += x
        sum_y += y
        sum_xx += x * x
        sum_yy += y * y
        sum_xy += x * y
    numerator = n * sum_xy - sum_x * sum_y
    denominator = math.sqrt(
        (n * sum_xx - sum_x * sum_x) * (n * sum_yy - sum_y * sum_y)
    )
    if denominator == 0:
        raise ValueError("Cannot calculate replicate correlation for a constant library")
    return numerator / denominator


def workbook_rows(source: Path) -> tuple[list[str], list[list[object]]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    if workbook.sheetnames != [EXPECTED_SHEET]:
        raise ValueError(
            f"Expected one worksheet named {EXPECTED_SHEET!r}; observed {workbook.sheetnames!r}"
        )
    worksheet = workbook[EXPECTED_SHEET]
    row_iterator = worksheet.iter_rows(values_only=True)
    header = [clean_text(value) for value in next(row_iterator)]
    expected_header = list(ANNOTATION_COLUMNS) + [spec[0] for spec in SAMPLE_SPECS]
    if header != expected_header:
        raise ValueError("Workbook columns do not match the frozen source schema")
    rows = [list(row) for row in row_iterator]
    workbook.close()
    return header, rows


def build_outputs(source: Path, output_dir: Path) -> None:
    observed_sha256 = sha256_file(source)
    if observed_sha256 != EXPECTED_SOURCE_SHA256:
        raise ValueError(
            "Source checksum mismatch: "
            f"expected {EXPECTED_SOURCE_SHA256}, observed {observed_sha256}"
        )

    header, source_rows = workbook_rows(source)
    header_index = {name: index for index, name in enumerate(header)}
    if len(source_rows) != 46_427:
        raise ValueError(f"Expected 46,427 gene rows; observed {len(source_rows):,}")

    records: list[tuple[int, list[str], list[int]]] = []
    observed_gene_ids: set[int] = set()
    symbol_members: dict[str, list[int]] = defaultdict(list)
    sample_library_sizes = Counter()
    sample_detected_genes = Counter()
    all_zero_genes = 0

    for row_number, row in enumerate(source_rows, start=2):
        gene_id_text = clean_text(row[header_index["Gene_ID"]])
        if not gene_id_text.isdigit():
            raise ValueError(f"Invalid Gene_ID at row {row_number}: {gene_id_text!r}")
        gene_id = int(gene_id_text)
        if gene_id in observed_gene_ids:
            raise ValueError(f"Duplicate Gene_ID at row {row_number}: {gene_id}")
        observed_gene_ids.add(gene_id)

        annotation = [clean_text(row[header_index[name]]) for name in ANNOTATION_COLUMNS]
        symbol = annotation[2]
        if not symbol:
            raise ValueError(f"Missing Gene_Symbol at row {row_number}")
        symbol_members[symbol].append(gene_id)

        counts = []
        for source_column, sample_id, *_ in SAMPLE_SPECS:
            count = parse_count(row[header_index[source_column]], row_number, source_column)
            counts.append(count)
            sample_library_sizes[sample_id] += count
            if count > 0:
                sample_detected_genes[sample_id] += 1
        if sum(counts) == 0:
            all_zero_genes += 1
        records.append((gene_id, annotation, counts))

    records.sort(key=lambda record: record[0])
    duplicate_symbols = {
        symbol: sorted(gene_ids)
        for symbol, gene_ids in symbol_members.items()
        if len(gene_ids) > 1
    }
    expected_duplicate = {
        "TRNAV-CAC": [107985614, 107985615, 107985753],
    }
    if duplicate_symbols != expected_duplicate:
        raise ValueError(
            f"Unexpected duplicate Gene_Symbol mapping: {duplicate_symbols!r}"
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    counts_path = output_dir / "gene_counts.tsv.gz"
    annotations_path = output_dir / "gene_annotations.tsv.gz"
    membership_path = output_dir / "gene_symbol_membership.tsv.gz"
    metadata_path = output_dir / "sample_metadata.tsv"

    write_tsv_gz(
        counts_path,
        ["gene_id", *[spec[1] for spec in SAMPLE_SPECS]],
        ([str(gene_id), *counts] for gene_id, _, counts in records),
    )
    annotation_header = (
        "gene_id",
        "transcript_id",
        "gene_symbol",
        "description",
        "gene_biotype",
        "protein_id",
        "hgnc_id",
        "mim_id",
        "ensembl_id",
        "imgt_gene_db_id",
    )
    write_tsv_gz(
        annotations_path,
        annotation_header,
        ([str(gene_id), *annotation[1:]] for gene_id, annotation, _ in records),
    )
    write_tsv_gz(
        membership_path,
        ["gene_symbol", "n_gene_ids", "gene_ids"],
        (
            [symbol, len(gene_ids), ";".join(map(str, sorted(gene_ids)))]
            for symbol, gene_ids in sorted(symbol_members.items())
        ),
    )
    write_tsv(
        metadata_path,
        [
            "sample_id",
            "source_column",
            "cell_line",
            "antigen_status",
            "genotype",
            "source_clone",
            "treatment",
            "tnf_ng_ml",
            "ifng_ng_ml",
            "treatment_duration_h",
            "replicate",
            "replicate_unit",
            "paired_batch_status",
        ],
        (
            [
                sample_id,
                source_column,
                "HeLa",
                "CD19_positive",
                genotype,
                clone,
                treatment,
                50 if treatment in {"TNF", "TNF_IFNG"} else 0,
                50 if treatment in {"IFNG", "TNF_IFNG"} else 0,
                48,
                replicate,
                "independent_experiment",
                "not_confirmed",
            ]
            for source_column, sample_id, genotype, clone, treatment, replicate in SAMPLE_SPECS
        ),
    )

    sample_index = {spec[1]: index for index, spec in enumerate(SAMPLE_SPECS)}
    grouped_samples: dict[str, list[str]] = defaultdict(list)
    for _, sample_id, genotype, _, treatment, _ in SAMPLE_SPECS:
        grouped_samples[f"{genotype}__{treatment}"].append(sample_id)
    within_group_correlations = {}
    for group, group_sample_ids in sorted(grouped_samples.items()):
        pairwise = []
        for sample_a, sample_b in combinations(group_sample_ids, 2):
            correlation = log1p_pearson(
                records, sample_index[sample_a], sample_index[sample_b]
            )
            pairwise.append(
                {
                    "sample_a": sample_a,
                    "sample_b": sample_b,
                    "pearson_r": correlation,
                }
            )
        correlations = [item["pearson_r"] for item in pairwise]
        within_group_correlations[group] = {
            "metric": "Pearson correlation over log1p raw counts for all Gene_ID rows",
            "minimum": min(correlations),
            "median": statistics.median(correlations),
            "maximum": max(correlations),
            "pairs": pairwise,
        }

    canonical_paths = [counts_path, annotations_path, membership_path, metadata_path]
    qc = {
        "source": {
            # The source identity is content-addressed; the frozen received
            # filename is provenance and does not depend on the local cache
            # basename used for a rebuild.
            "filename": SOURCE_FILENAME_AS_RECEIVED,
            "sha256": observed_sha256,
            "size_bytes": source.stat().st_size,
            "worksheet": EXPECTED_SHEET,
            "rows_including_header": len(source_rows) + 1,
            "columns": len(header),
        },
        "count_matrix": {
            "gene_id_rows": len(records),
            "unique_gene_ids": len(observed_gene_ids),
            "unique_gene_symbols": len(symbol_members),
            "samples": len(SAMPLE_SPECS),
            "all_zero_gene_ids_across_all_samples": all_zero_genes,
            "missing_counts": 0,
            "non_integer_counts": 0,
            "negative_counts": 0,
        },
        "duplicate_gene_symbols": duplicate_symbols,
        "gene_symbol_aggregation": (
            "For symbol-level analyses, counts are summed over Gene_ID rows before "
            "model fitting. This affects only TRNAV-CAC in this source matrix."
        ),
        "sample_qc": {
            sample_id: {
                "library_size": sample_library_sizes[sample_id],
                "detected_gene_ids": sample_detected_genes[sample_id],
            }
            for _, sample_id, *_ in SAMPLE_SPECS
        },
        "within_group_replicate_correlations": within_group_correlations,
        "canonical_files": {
            path.name: {
                "size_bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
            for path in canonical_paths
        },
    }
    qc_path = output_dir / "source_qc.json"
    qc_path.write_text(
        json.dumps(qc, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    checksum_paths = [*canonical_paths, qc_path]
    checksum_lines = [
        f"{sha256_file(path)}  {path.name}" for path in sorted(checksum_paths)
    ]
    (output_dir / "SHA256SUMS").write_text(
        "\n".join(checksum_lines) + "\n", encoding="utf-8"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Macrogen XLSX workbook")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data/experimental/bulk_rnaseq"),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build_outputs(args.input, args.output_dir)


if __name__ == "__main__":
    main()
