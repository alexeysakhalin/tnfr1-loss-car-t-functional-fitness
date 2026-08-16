#!/usr/bin/env python3
"""Validate the OOXML structure and marker-table layout of an XLSX workbook."""

from __future__ import annotations

import argparse
import posixpath
import re
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from urllib.parse import unquote, urlsplit
from xml.etree import ElementTree as ET


PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
DOCUMENT_REL_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_DOCUMENT_REL_SUFFIX = "/officeDocument"
WORKSHEET_REL_SUFFIX = "/worksheet"

CELL_REF_RE = re.compile(r"^\$?([A-Z]{1,3})\$?([1-9][0-9]*)$")
DIMENSION_RE = re.compile(
    r"^(\$?[A-Z]{1,3}\$?[1-9][0-9]*)(?::(\$?[A-Z]{1,3}\$?[1-9][0-9]*))?$"
)


class WorkbookValidationError(ValueError):
    """Raised when an XLSX package violates the release contract."""


@dataclass(frozen=True)
class WorksheetExtent:
    dimension: str
    min_row: int
    min_column: int
    max_row: int
    max_column: int


def _column_number(label: str) -> int:
    result = 0
    for character in label:
        result = result * 26 + ord(character) - ord("A") + 1
    return result


def _column_label(number: int) -> str:
    if number < 1:
        raise ValueError("Column numbers are one-based.")
    characters: list[str] = []
    while number:
        number, remainder = divmod(number - 1, 26)
        characters.append(chr(ord("A") + remainder))
    return "".join(reversed(characters))


def _cell_coordinates(reference: str) -> tuple[int, int]:
    match = CELL_REF_RE.fullmatch(reference)
    if match is None:
        raise WorkbookValidationError(f"Invalid cell reference: {reference!r}")
    return int(match.group(2)), _column_number(match.group(1))


def _parse_xml(archive: zipfile.ZipFile, part: str) -> ET.Element:
    try:
        data = archive.read(part)
    except KeyError as exc:
        raise WorkbookValidationError(f"Missing required OOXML part: {part}") from exc
    try:
        return ET.fromstring(data)
    except ET.ParseError as exc:
        raise WorkbookValidationError(f"Malformed XML in {part}: {exc}") from exc


def _relationship_source(relationship_part: str) -> str:
    path = PurePosixPath(relationship_part)
    if relationship_part == "_rels/.rels":
        return ""
    if path.parent.name != "_rels" or not path.name.endswith(".rels"):
        raise WorkbookValidationError(
            f"Invalid relationship-part location: {relationship_part}"
        )
    return str(path.parent.parent / path.name[: -len(".rels")])


def _resolve_relationship_target(source: str, target: str) -> str:
    parsed = urlsplit(target)
    if parsed.scheme or parsed.netloc:
        raise WorkbookValidationError(
            f"Internal relationship uses an absolute URI: {target!r}"
        )
    target_path = unquote(parsed.path).replace("\\", "/")
    if target_path.startswith("/"):
        resolved = posixpath.normpath(target_path.lstrip("/"))
    else:
        resolved = posixpath.normpath(
            posixpath.join(posixpath.dirname(source), target_path)
        )
    if resolved in ("", ".", "..") or resolved.startswith("../"):
        raise WorkbookValidationError(
            f"Relationship target escapes the XLSX package: {target!r}"
        )
    return resolved


def _relationship_map(
    archive: zipfile.ZipFile,
    relationship_part: str,
    members: set[str],
) -> dict[str, tuple[str, str]]:
    root = _parse_xml(archive, relationship_part)
    source = _relationship_source(relationship_part)
    relationships: dict[str, tuple[str, str]] = {}
    for relationship in root.findall(f"{{{PACKAGE_REL_NS}}}Relationship"):
        relationship_id = relationship.get("Id")
        relationship_type = relationship.get("Type")
        target = relationship.get("Target")
        if not relationship_id or not relationship_type or not target:
            raise WorkbookValidationError(
                f"Incomplete relationship in {relationship_part}"
            )
        if relationship_id in relationships:
            raise WorkbookValidationError(
                f"Duplicate relationship Id {relationship_id!r} in {relationship_part}"
            )
        if relationship.get("TargetMode") == "External":
            relationships[relationship_id] = (relationship_type, target)
            continue
        resolved_target = _resolve_relationship_target(source, target)
        if resolved_target not in members:
            raise WorkbookValidationError(
                f"{relationship_part} points to missing part {resolved_target!r} "
                f"({relationship_id})"
            )
        relationships[relationship_id] = (relationship_type, resolved_target)
    return relationships


def _validate_all_relationships(
    archive: zipfile.ZipFile, members: set[str]
) -> dict[str, tuple[str, str]]:
    relationship_parts = sorted(
        member for member in members if member.endswith(".rels")
    )
    if "_rels/.rels" not in relationship_parts:
        raise WorkbookValidationError("Missing package relationships: _rels/.rels")
    root_relationships: dict[str, tuple[str, str]] | None = None
    for part in relationship_parts:
        relationships = _relationship_map(archive, part, members)
        if part == "_rels/.rels":
            root_relationships = relationships
    assert root_relationships is not None
    return root_relationships


def _validate_content_types(archive: zipfile.ZipFile, members: set[str]) -> None:
    root = _parse_xml(archive, "[Content_Types].xml")
    if root.tag != f"{{{CONTENT_TYPES_NS}}}Types":
        raise WorkbookValidationError("[Content_Types].xml has an invalid root element.")
    seen_parts: set[str] = set()
    for override in root.findall(f"{{{CONTENT_TYPES_NS}}}Override"):
        part_name = override.get("PartName")
        content_type = override.get("ContentType")
        if not part_name or not content_type or not part_name.startswith("/"):
            raise WorkbookValidationError("Invalid content-type Override entry.")
        part = unquote(part_name.lstrip("/"))
        if part in seen_parts:
            raise WorkbookValidationError(
                f"Duplicate content-type Override for {part!r}."
            )
        seen_parts.add(part)
        if part not in members:
            raise WorkbookValidationError(
                f"[Content_Types].xml declares missing part {part!r}."
            )


def _workbook_part(root_relationships: dict[str, tuple[str, str]]) -> str:
    candidates = [
        target
        for relationship_type, target in root_relationships.values()
        if relationship_type.endswith(OFFICE_DOCUMENT_REL_SUFFIX)
        and not target.startswith(("http://", "https://"))
    ]
    if len(candidates) != 1:
        raise WorkbookValidationError(
            "Expected exactly one internal officeDocument relationship; "
            f"found {len(candidates)}"
        )
    return candidates[0]


def _workbook_relationship_part(workbook_part: str) -> str:
    path = PurePosixPath(workbook_part)
    return str(path.parent / "_rels" / f"{path.name}.rels")


def _sheet_parts(
    archive: zipfile.ZipFile,
    workbook_part: str,
    members: set[str],
) -> list[tuple[str, str]]:
    workbook = _parse_xml(archive, workbook_part)
    relationship_part = _workbook_relationship_part(workbook_part)
    relationships = _relationship_map(archive, relationship_part, members)
    sheets = workbook.find(f"{{{SPREADSHEET_NS}}}sheets")
    if sheets is None:
        raise WorkbookValidationError("Workbook has no <sheets> collection.")

    result: list[tuple[str, str]] = []
    seen_names: set[str] = set()
    for sheet in sheets.findall(f"{{{SPREADSHEET_NS}}}sheet"):
        name = sheet.get("name")
        relationship_id = sheet.get(f"{{{DOCUMENT_REL_NS}}}id")
        if not name or not relationship_id:
            raise WorkbookValidationError("Workbook contains an unnamed/unlinked sheet.")
        if name in seen_names:
            raise WorkbookValidationError(f"Duplicate worksheet name: {name!r}")
        seen_names.add(name)
        try:
            relationship_type, target = relationships[relationship_id]
        except KeyError as exc:
            raise WorkbookValidationError(
                f"Worksheet {name!r} has unresolved relationship {relationship_id!r}"
            ) from exc
        if not relationship_type.endswith(WORKSHEET_REL_SUFFIX):
            raise WorkbookValidationError(
                f"Worksheet {name!r} relationship is not a worksheet part."
            )
        result.append((name, target))
    if not result:
        raise WorkbookValidationError("Workbook contains no worksheets.")
    return result


def _worksheet_extent(archive: zipfile.ZipFile, part: str) -> WorksheetExtent:
    worksheet = _parse_xml(archive, part)
    dimension_node = worksheet.find(f"{{{SPREADSHEET_NS}}}dimension")
    dimension = dimension_node.get("ref") if dimension_node is not None else None
    if not dimension or DIMENSION_RE.fullmatch(dimension) is None:
        raise WorkbookValidationError(
            f"Worksheet {part} has an invalid or missing dimension: {dimension!r}"
        )

    coordinates: list[tuple[int, int]] = []
    references: set[str] = set()
    for cell in worksheet.findall(
        f".//{{{SPREADSHEET_NS}}}sheetData/{{{SPREADSHEET_NS}}}row/"
        f"{{{SPREADSHEET_NS}}}c"
    ):
        reference = cell.get("r")
        if not reference:
            raise WorkbookValidationError(f"Worksheet {part} contains an unaddressed cell.")
        normalized_reference = reference.replace("$", "")
        if normalized_reference in references:
            raise WorkbookValidationError(
                f"Worksheet {part} contains duplicate cell {normalized_reference}."
            )
        references.add(normalized_reference)
        coordinates.append(_cell_coordinates(reference))
    if not coordinates:
        raise WorkbookValidationError(f"Worksheet {part} contains no cells.")

    min_row = min(row for row, _ in coordinates)
    min_column = min(column for _, column in coordinates)
    max_row = max(row for row, _ in coordinates)
    max_column = max(column for _, column in coordinates)
    expected_dimension = (
        f"{_column_label(min_column)}{min_row}:"
        f"{_column_label(max_column)}{max_row}"
    )
    if min_row == max_row and min_column == max_column:
        expected_dimension = f"{_column_label(min_column)}{min_row}"
    if dimension.replace("$", "") != expected_dimension:
        raise WorkbookValidationError(
            f"Worksheet {part} declares dimension {dimension!r}, but populated "
            f"cells require {expected_dimension!r}."
        )
    return WorksheetExtent(
        dimension=dimension,
        min_row=min_row,
        min_column=min_column,
        max_row=max_row,
        max_column=max_column,
    )


def _parse_clusters(value: str) -> list[str]:
    clusters = [item.strip() for item in value.split(",") if item.strip()]
    if not clusters or len(clusters) != len(set(clusters)):
        raise argparse.ArgumentTypeError(
            "--expected-clusters must be a comma-separated list of unique labels"
        )
    return clusters


def verify_marker_workbook(
    path: Path,
    *,
    expected_clusters: list[str],
    top_n: int,
    expected_columns: int,
    require_openpyxl: bool = False,
) -> None:
    """Validate an R/05 per-cluster marker workbook."""
    if top_n < 1 or expected_columns < 1:
        raise WorkbookValidationError("top_n and expected_columns must be positive.")
    try:
        with zipfile.ZipFile(path) as archive:
            corrupt_member = archive.testzip()
            if corrupt_member is not None:
                raise WorkbookValidationError(
                    f"ZIP CRC validation failed for {corrupt_member!r}."
                )
            names = archive.namelist()
            if len(names) != len(set(names)):
                raise WorkbookValidationError("XLSX contains duplicate ZIP members.")
            if any(
                name.startswith("/")
                or "\\" in name
                or ".." in PurePosixPath(name).parts
                for name in names
            ):
                raise WorkbookValidationError("XLSX contains a non-canonical ZIP path.")
            members = set(names)
            for required in ("[Content_Types].xml", "_rels/.rels"):
                if required not in members:
                    raise WorkbookValidationError(
                        f"Missing required OOXML part: {required}"
                    )

            _validate_content_types(archive, members)
            root_relationships = _validate_all_relationships(archive, members)
            workbook_part = _workbook_part(root_relationships)
            sheets = _sheet_parts(archive, workbook_part, members)
            expected_names = ["All_clusters_top20"] + [
                f"Cluster_{cluster}" for cluster in expected_clusters
            ]
            observed_names = [name for name, _ in sheets]
            if observed_names != expected_names:
                raise WorkbookValidationError(
                    "Worksheet order/name contract failed: expected "
                    f"{expected_names!r}, observed {observed_names!r}."
                )

            extents = {
                name: _worksheet_extent(archive, part) for name, part in sheets
            }
    except (OSError, zipfile.BadZipFile) as exc:
        raise WorkbookValidationError(f"Cannot read XLSX {path}: {exc}") from exc

    expected_summary_rows = 1 + top_n * len(expected_clusters)
    for name, extent in extents.items():
        expected_rows = (
            expected_summary_rows if name == "All_clusters_top20" else 1 + top_n
        )
        if (extent.min_row, extent.min_column) != (1, 1):
            raise WorkbookValidationError(
                f"Worksheet {name!r} must begin at A1; observed {extent.dimension}."
            )
        if extent.max_row != expected_rows or extent.max_column != expected_columns:
            raise WorkbookValidationError(
                f"Worksheet {name!r} extent contract failed: expected "
                f"A1:{_column_label(expected_columns)}{expected_rows}, observed "
                f"{extent.dimension}."
            )

    if require_openpyxl:
        try:
            import openpyxl
        except ImportError as exc:
            raise WorkbookValidationError(
                "openpyxl is required for the requested reader validation."
            ) from exc
        try:
            workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
        except Exception as exc:
            raise WorkbookValidationError(
                f"openpyxl could not load {path}: {type(exc).__name__}: {exc}"
            ) from exc
        try:
            if workbook.sheetnames != expected_names:
                raise WorkbookValidationError(
                    "openpyxl observed an unexpected worksheet sequence: "
                    f"{workbook.sheetnames!r}."
                )
            for name, extent in extents.items():
                worksheet = workbook[name]
                if (
                    worksheet.max_row != extent.max_row
                    or worksheet.max_column != extent.max_column
                ):
                    raise WorkbookValidationError(
                        f"openpyxl extent mismatch for {name!r}: "
                        f"{worksheet.max_row} rows x {worksheet.max_column} columns."
                    )
                header_is_incomplete = any(
                    worksheet.cell(1, column).value is None
                    for column in range(1, expected_columns + 1)
                )
                if header_is_incomplete:
                    raise WorkbookValidationError(
                        f"openpyxl observed a blank header in worksheet {name!r}."
                    )
        finally:
            workbook.close()


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--expected-clusters",
        required=True,
        type=_parse_clusters,
        help="Comma-separated cluster labels in required worksheet order.",
    )
    parser.add_argument("--top-n", required=True, type=int)
    parser.add_argument("--expected-columns", required=True, type=int)
    parser.add_argument(
        "--require-openpyxl",
        action="store_true",
        help="Also require a successful normal-mode openpyxl load.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        verify_marker_workbook(
            args.workbook,
            expected_clusters=args.expected_clusters,
            top_n=args.top_n,
            expected_columns=args.expected_columns,
            require_openpyxl=args.require_openpyxl,
        )
    except WorkbookValidationError as exc:
        print(f"XLSX validation failed: {exc}", file=sys.stderr)
        return 1
    print(f"XLSX validation passed: {args.workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
