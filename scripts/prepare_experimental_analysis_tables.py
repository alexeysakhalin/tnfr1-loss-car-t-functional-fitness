#!/usr/bin/env python3
"""Validate project workbooks and create version-controlled analysis tables.

The source Excel workbooks remain unchanged.  This script converts only the
processed, non-identifying values needed by the figure scripts into stable,
compressed TSV files.  It also records source and output checksums so that a
reader can verify the exact transformation.
"""

from __future__ import annotations

import argparse
import csv
import fcntl
import gzip
import hashlib
import json
import math
import os
import statistics
import tempfile
from collections import Counter
from pathlib import Path
from typing import Iterable, Iterator, Sequence

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "data" / "experimental"

BULK_COLUMNS = (
    "Gene_Symbol",
    "baseMean",
    "log2FoldChange",
    "lfcSE",
    "stat",
    "pvalue",
    "padj",
)

MATCHED_COLUMNS = (
    "treatment",
    "Gene_Symbol",
    "baseMean",
    "log2FoldChange",
    "lfcSE",
    "stat",
    "pvalue",
    "padj",
)

SOURCE_SPECS = {
    "TNF": {
        "pattern": "Differential_Expression_TNFa_vs_control*.xlsx",
        "canonical_name": "Differential_Expression_TNFa_vs_control_final_filtered.xlsx",
    },
    "IFNG": {
        "pattern": "Differential_Expression_IFN_vs_control*.xlsx",
        "canonical_name": "Differential_Expression_IFN_vs_control_filtered.xlsx",
    },
    "TNF_IFNG": {
        "pattern": "Differential_Expression_TI_vs_control*.xlsx",
        "canonical_name": "Differential_Expression_TI_vs_control_final_filtered.xlsx",
    },
    "T6_MATCHED": {
        "pattern": "PyDESeq2_T6_vs_Hela_matched_treatments*.xlsx",
        "canonical_name": "PyDESeq2_T6_vs_Hela_matched_treatments.xlsx",
    },
    "WT": {"pattern": "WT_new*.xlsx", "canonical_name": "WT_new.xlsx"},
    "KO1": {"pattern": "KO1_new*.xlsx", "canonical_name": "KO1_new.xlsx"},
    "KO2": {"pattern": "KO2_new*.xlsx", "canonical_name": "KO2_new.xlsx"},
    "TCR": {"pattern": "TCR.xlsx", "canonical_name": "TCR.xlsx"},
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def locate_one(input_dir: Path, pattern: str) -> Path:
    matches = sorted(input_dir.glob(pattern))
    if len(matches) != 1:
        raise RuntimeError(
            f"Expected exactly one file matching {pattern!r} in {input_dir}; "
            f"found {len(matches)}: {[path.name for path in matches]}"
        )
    return matches[0]


def finite_number(value: object, label: str, *, allow_none: bool = False) -> float | None:
    if value is None and allow_none:
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{label} is not numeric: {value!r}")
    number = float(value)
    if not math.isfinite(number):
        raise ValueError(f"{label} is not finite: {value!r}")
    return number


def clean_gene(value: object, label: str) -> str:
    gene = str(value).strip() if value is not None else ""
    if not gene:
        raise ValueError(f"Blank gene symbol in {label}")
    return gene.upper()


def format_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return format(value, ".17g")
    return str(value)


def write_tsv(path: Path, header: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        with temporary_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
            writer.writerow(header)
            for row in rows:
                writer.writerow([format_value(value) for value in row])
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def write_tsv_gzip(
    path: Path,
    header: Sequence[str],
    rows: Iterable[Sequence[object]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        with temporary_path.open("wb") as raw:
            with gzip.GzipFile(
                filename="", mode="wb", fileobj=raw, mtime=0
            ) as compressed:
                with __import__("io").TextIOWrapper(
                    compressed, encoding="utf-8", newline=""
                ) as handle:
                    writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
                    writer.writerow(header)
                    for row in rows:
                        writer.writerow([format_value(value) for value in row])
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def write_text_atomic(path: Path, text_value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        temporary_path.write_text(text_value, encoding="utf-8")
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def update_row_digest(digest: object, row: Sequence[object]) -> None:
    payload = json.dumps(
        list(row), ensure_ascii=False, separators=(",", ":"),
        allow_nan=False, default=str
    )
    digest.update(payload.encode("utf-8"))
    digest.update(b"\n")


def acquire_lock(lock_path: Path):
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    lock_handle = lock_path.open("a+", encoding="utf-8")
    try:
        fcntl.flock(lock_handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError as error:
        lock_handle.close()
        raise RuntimeError(
            f"Another experimental-data preparation run holds {lock_path}"
        ) from error
    lock_handle.seek(0)
    lock_handle.truncate()
    lock_handle.write(f"pid={os.getpid()}\n")
    lock_handle.flush()
    return lock_handle


def remove_stale_temporary_files(output_dir: Path) -> None:
    if not output_dir.exists():
        return
    for path in output_dir.rglob(".*.tmp"):
        if path.is_file():
            path.unlink()


def workbook_rows(path: Path, sheet_name: str) -> tuple[list[str], Iterator[tuple[object, ...]], object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(
            f"Sheet {sheet_name!r} is absent from {path.name}; "
            f"available sheets: {workbook.sheetnames}"
        )
    worksheet = workbook[sheet_name]
    iterator = worksheet.iter_rows(values_only=True)
    header = [str(value).strip() if value is not None else "" for value in next(iterator)]
    return header, iterator, workbook


def prepare_cytokine_de(
    sources: dict[str, Path], output_dir: Path
) -> tuple[dict[str, object], list[dict[str, object]]]:
    output_path = output_dir / "hela_cytokine_significant_differential_expression.tsv.gz"
    (output_dir / "hela_cytokine_differential_expression.tsv.gz").unlink(
        missing_ok=True
    )
    output_rows: list[tuple[object, ...]] = []
    condition_qc: dict[str, object] = {}

    for condition in ("TNF", "IFNG", "TNF_IFNG"):
        path = sources[condition]
        header, iterator, workbook = workbook_rows(path, "Sheet1")
        if tuple(header) != BULK_COLUMNS:
            workbook.close()
            raise ValueError(f"Unexpected columns in {path.name}: {header}")
        index = {name: position for position, name in enumerate(header)}
        seen: set[str] = set()
        source_sign_reversed = 0
        zero_p = 0
        zero_padj = 0
        non_significant = 0
        n_rows = 0

        for row_number, row in enumerate(iterator, start=2):
            n_rows += 1
            gene = clean_gene(row[index["Gene_Symbol"]], f"{path.name}:{row_number}")
            if gene in seen:
                workbook.close()
                raise ValueError(f"Duplicate gene {gene} in {path.name}")
            seen.add(gene)

            base_mean = finite_number(row[index["baseMean"]], f"{path.name}:{gene}:baseMean")
            log2_fc = finite_number(
                row[index["log2FoldChange"]], f"{path.name}:{gene}:log2FoldChange"
            )
            lfc_se = finite_number(row[index["lfcSE"]], f"{path.name}:{gene}:lfcSE")
            source_stat = finite_number(row[index["stat"]], f"{path.name}:{gene}:stat")
            p_value = finite_number(row[index["pvalue"]], f"{path.name}:{gene}:pvalue")
            adjusted_p = finite_number(row[index["padj"]], f"{path.name}:{gene}:padj")

            assert base_mean is not None and log2_fc is not None and lfc_se is not None
            assert source_stat is not None and p_value is not None and adjusted_p is not None
            if base_mean < 0 or lfc_se <= 0:
                workbook.close()
                raise ValueError(f"Invalid base mean or LFC SE for {gene} in {path.name}")
            if not 0 <= p_value <= 1 or not 0 <= adjusted_p <= 1:
                workbook.close()
                raise ValueError(f"Invalid p value for {gene} in {path.name}")

            canonical_stat = log2_fc / lfc_se
            if not math.isclose(abs(canonical_stat), abs(source_stat), rel_tol=1e-10, abs_tol=1e-10):
                workbook.close()
                raise ValueError(
                    f"Wald statistic magnitude mismatch for {gene} in {path.name}"
                )
            if canonical_stat != 0 and source_stat != 0 and math.copysign(1, canonical_stat) != math.copysign(1, source_stat):
                source_sign_reversed += 1
            if p_value == 0:
                zero_p += 1
            if adjusted_p == 0:
                zero_padj += 1
            if adjusted_p >= 0.05:
                non_significant += 1

            output_rows.append(
                (
                    condition,
                    gene,
                    base_mean,
                    log2_fc,
                    lfc_se,
                    canonical_stat,
                    p_value,
                    adjusted_p,
                )
            )
        workbook.close()
        if n_rows == 0 or len(seen) != n_rows:
            raise ValueError(f"Empty or non-unique cytokine table: {path.name}")
        if non_significant != 0:
            raise ValueError(
                f"{path.name} is not an FDR-significant-only table: "
                f"{non_significant} row(s) have adjusted p >= 0.05"
            )
        condition_qc[condition] = {
            "rows": n_rows,
            "unique_genes": len(seen),
            "source_stat_sign_reversed_rows": source_sign_reversed,
            "zero_p_values": zero_p,
            "zero_adjusted_p_values": zero_padj,
            "rows_with_adjusted_p_at_least_0_05": non_significant,
        }

    write_tsv_gzip(
        output_path,
        (
            "condition",
            "gene_symbol",
            "base_mean",
            "log2_fold_change_treatment_vs_untreated",
            "lfc_se",
            "wald_statistic_treatment_vs_untreated",
            "p_value",
            "adjusted_p_value",
        ),
        output_rows,
    )
    return {
        "path": output_path.relative_to(ROOT).as_posix(),
        "sha256": sha256(output_path),
        "size_bytes": output_path.stat().st_size,
        "rows": len(output_rows),
    }, [condition_qc]


def prepare_t6_matched(
    source: Path, output_dir: Path
) -> tuple[dict[str, object], list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    required = ["SampleSheet", "UNTREATED", "IFN", "TNF", "TI", "ALL"]
    if workbook.sheetnames != required:
        workbook.close()
        raise ValueError(
            f"Unexpected sheets in {source.name}: {workbook.sheetnames}; expected {required}"
        )

    sample_sheet = workbook["SampleSheet"]
    sample_rows = sample_sheet.iter_rows(values_only=True)
    sample_header = [str(value).strip() for value in next(sample_rows)]
    if sample_header != ["sample", "cell_line", "treatment", "replicate"]:
        workbook.close()
        raise ValueError(f"Unexpected SampleSheet columns: {sample_header}")
    design: list[tuple[object, ...]] = []
    sample_ids: set[str] = set()
    design_counter: Counter[tuple[str, str]] = Counter()
    replicate_sets: dict[tuple[str, str], set[int]] = {}
    for row in sample_rows:
        sample, cell_line, treatment, replicate = row
        sample = str(sample).strip()
        cell_line = str(cell_line).strip()
        treatment = str(treatment).strip()
        replicate = int(replicate)
        if sample in sample_ids:
            workbook.close()
            raise ValueError(f"Duplicate sample in SampleSheet: {sample}")
        sample_ids.add(sample)
        key = (cell_line, treatment)
        design_counter[key] += 1
        replicate_sets.setdefault(key, set()).add(replicate)
        design.append((sample, cell_line, treatment, replicate))
    expected_keys = {
        (cell_line, treatment)
        for cell_line in ("Hela", "T6")
        for treatment in ("UNTREATED", "IFN", "TNF", "TI")
    }
    if set(design_counter) != expected_keys:
        workbook.close()
        raise ValueError(f"Incomplete T6 design: {dict(design_counter)}")
    for key in expected_keys:
        if design_counter[key] != 3 or replicate_sets[key] != {1, 2, 3}:
            workbook.close()
            raise ValueError(f"Unbalanced T6 design at {key}: n={design_counter[key]}, reps={replicate_sets[key]}")

    output_rows_by_sheet: dict[str, list[tuple[object, ...]]] = {}
    reference_genes: set[str] | None = None
    sheet_qc: dict[str, object] = {}
    condition_source_digest = hashlib.sha256()
    condition_source_rows = 0
    for sheet_name in ("UNTREATED", "IFN", "TNF", "TI"):
        sheet_output_rows: list[tuple[object, ...]] = []
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = [str(value).strip() if value is not None else "" for value in next(rows)]
        if tuple(header) != MATCHED_COLUMNS:
            workbook.close()
            raise ValueError(f"Unexpected columns in {source.name}:{sheet_name}: {header}")
        index = {name: position for position, name in enumerate(header)}
        genes: set[str] = set()
        n_rows = 0
        n_adjusted_missing = 0
        for row_number, row in enumerate(rows, start=2):
            n_rows += 1
            condition_source_rows += 1
            update_row_digest(condition_source_digest, row)
            treatment = str(row[index["treatment"]]).strip()
            if treatment != sheet_name:
                workbook.close()
                raise ValueError(f"Treatment mismatch in {sheet_name}:{row_number}")
            gene = clean_gene(row[index["Gene_Symbol"]], f"{sheet_name}:{row_number}")
            if gene in genes:
                workbook.close()
                raise ValueError(f"Duplicate gene {gene} in {sheet_name}")
            genes.add(gene)

            values = [
                finite_number(
                    row[index[column]],
                    f"{source.name}:{sheet_name}:{gene}:{column}",
                    allow_none=(column != "baseMean"),
                )
                for column in ("baseMean", "log2FoldChange", "lfcSE", "stat", "pvalue", "padj")
            ]
            base_mean, log2_fc, lfc_se, statistic, p_value, adjusted_p = values
            assert base_mean is not None
            if base_mean < 0:
                workbook.close()
                raise ValueError(f"Negative base mean for {gene} in {sheet_name}")
            effect_test = [log2_fc, lfc_se, statistic]
            if any(value is None for value in effect_test) and not all(
                value is None for value in effect_test
            ):
                workbook.close()
                raise ValueError(f"Partially missing test result for {gene} in {sheet_name}")
            if log2_fc is not None:
                assert lfc_se is not None and statistic is not None
                if lfc_se <= 0 or not math.isclose(
                    log2_fc / lfc_se, statistic, rel_tol=1e-10, abs_tol=1e-10
                ):
                    workbook.close()
                    raise ValueError(f"Invalid Wald statistic for {gene} in {sheet_name}")
                if p_value is not None and not 0 <= p_value <= 1:
                    workbook.close()
                    raise ValueError(f"Invalid p value for {gene} in {sheet_name}")
            if adjusted_p is None:
                n_adjusted_missing += 1
            elif not 0 <= adjusted_p <= 1:
                workbook.close()
                raise ValueError(f"Invalid adjusted p value for {gene} in {sheet_name}")

            sheet_output_rows.append(
                (
                    sheet_name,
                    gene,
                    base_mean,
                    log2_fc,
                    lfc_se,
                    statistic,
                    p_value,
                    adjusted_p,
                )
            )
        if reference_genes is None:
            reference_genes = genes
        elif genes != reference_genes:
            workbook.close()
            raise ValueError(f"Gene universe differs in {sheet_name}")
        sheet_qc[sheet_name] = {
            "rows": n_rows,
            "unique_genes": len(genes),
            "adjusted_p_missing": n_adjusted_missing,
        }
        output_rows_by_sheet[sheet_name] = sheet_output_rows

    all_worksheet = workbook["ALL"]
    all_rows = all_worksheet.iter_rows(values_only=True)
    all_header = [
        str(value).strip() if value is not None else "" for value in next(all_rows)
    ]
    if tuple(all_header) != MATCHED_COLUMNS:
        workbook.close()
        raise ValueError(f"Unexpected columns in {source.name}:ALL: {all_header}")
    all_source_digest = hashlib.sha256()
    all_source_rows = 0
    for row in all_rows:
        all_source_rows += 1
        update_row_digest(all_source_digest, row)
    if (
        all_source_rows != condition_source_rows
        or all_source_digest.hexdigest() != condition_source_digest.hexdigest()
    ):
        workbook.close()
        raise ValueError(
            "The ALL sheet is not the ordered union of UNTREATED, IFN, TNF and TI."
        )
    workbook.close()

    design_path = output_dir / "hela_t6_matched_sample_design.tsv"
    write_tsv(design_path, sample_header, design)
    legacy_combined = output_dir / "hela_t6_vs_wt_matched_differential_expression.tsv.gz"
    legacy_combined.unlink(missing_ok=True)
    de_outputs: list[dict[str, object]] = []
    for sheet_name in ("UNTREATED", "IFN", "TNF", "TI"):
        de_path = output_dir / (
            "hela_t6_vs_wt_" + sheet_name.lower() + "_differential_expression.tsv.gz"
        )
        write_tsv_gzip(
            de_path,
            (
                "condition",
                "gene_symbol",
                "base_mean",
                "log2_fold_change_t6_vs_wt",
                "lfc_se",
                "wald_statistic_t6_vs_wt",
                "p_value",
                "adjusted_p_value",
            ),
            output_rows_by_sheet[sheet_name],
        )
        de_outputs.append(
            {
                "path": de_path.relative_to(ROOT).as_posix(),
                "sha256": sha256(de_path),
                "size_bytes": de_path.stat().st_size,
                "rows": len(output_rows_by_sheet[sheet_name]),
            }
        )
    return (
        {
            "path": design_path.relative_to(ROOT).as_posix(),
            "sha256": sha256(design_path),
            "size_bytes": design_path.stat().st_size,
            "rows": len(design),
        },
        de_outputs,
        {
            "design": {
                f"{cell_line}|{treatment}": design_counter[(cell_line, treatment)]
                for cell_line, treatment in sorted(design_counter)
            },
            "sheets": sheet_qc,
            "all_sheet": {
                "rows": all_source_rows,
                "matches_ordered_condition_union": True,
                "source_row_digest_sha256": all_source_digest.hexdigest(),
            },
        },
    )


def prepare_single_cell(
    sources: dict[str, Path], output_dir: Path
) -> tuple[list[dict[str, object]], dict[str, object]]:
    singlecell_dir = output_dir / "singlecell"
    outputs: list[dict[str, object]] = []
    qc: dict[str, object] = {}
    reference_genes: list[str] | None = None

    for sample in ("WT", "KO1", "KO2", "TCR"):
        path = sources[sample]
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = "CD3+ cells"
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet {sheet_name!r} absent from {path.name}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = [str(value).strip() if value is not None else "" for value in next(rows)]
        if header[0] != "Cell_Index" or len(header) < 3:
            workbook.close()
            raise ValueError(f"Unexpected single-cell header in {path.name}")
        genes = [clean_gene(value, f"{path.name}:header") for value in header[1:]]
        if len(genes) != len(set(genes)):
            workbook.close()
            raise ValueError(f"Duplicate gene columns in {path.name}")
        if reference_genes is None:
            reference_genes = genes
        elif genes != reference_genes:
            workbook.close()
            raise ValueError(f"Gene panel differs in {path.name}")

        output_rows: list[tuple[object, ...]] = []
        cell_ids: set[str] = set()
        library_sizes: list[int] = []
        detected_features: list[int] = []
        for row_number, row in enumerate(rows, start=2):
            cell_index = str(row[0]).strip() if row[0] is not None else ""
            if not cell_index:
                workbook.close()
                raise ValueError(f"Blank cell index in {path.name}:{row_number}")
            if cell_index in cell_ids:
                workbook.close()
                raise ValueError(f"Duplicate cell index {cell_index} in {path.name}")
            cell_ids.add(cell_index)

            counts: list[int] = []
            for column_number, value in enumerate(row[1:], start=2):
                number = finite_number(
                    value, f"{path.name}:{row_number}:{column_number}"
                )
                assert number is not None
                if number < 0 or not number.is_integer():
                    workbook.close()
                    raise ValueError(
                        f"Non-negative integer count expected in {path.name}:"
                        f"{row_number}:{column_number}; observed {number}"
                    )
                counts.append(int(number))
            if len(counts) != len(genes):
                workbook.close()
                raise ValueError(f"Row width mismatch in {path.name}:{row_number}")
            library_size = sum(counts)
            if library_size <= 0:
                workbook.close()
                raise ValueError(f"Zero-count cell in {path.name}:{row_number}")
            library_sizes.append(library_size)
            detected_features.append(sum(value > 0 for value in counts))
            output_rows.append((cell_index, *counts))
        workbook.close()

        output_path = singlecell_dir / f"{sample}_targeted_counts.tsv.gz"
        write_tsv_gzip(output_path, ("cell_index", *genes), output_rows)
        outputs.append(
            {
                "path": output_path.relative_to(ROOT).as_posix(),
                "sha256": sha256(output_path),
                "size_bytes": output_path.stat().st_size,
                "rows": len(output_rows),
            }
        )
        qc[sample] = {
            "source_sheet": sheet_name,
            "cells": len(output_rows),
            "genes": len(genes),
            "library_size_min": min(library_sizes),
            "library_size_median": statistics.median(library_sizes),
            "library_size_max": max(library_sizes),
            "detected_features_min": min(detected_features),
            "detected_features_median": statistics.median(detected_features),
            "detected_features_max": max(detected_features),
        }
    return outputs, qc


def write_manifest(
    sources: dict[str, Path], outputs: list[dict[str, object]], output_dir: Path
) -> Path:
    manifest_path = output_dir / "experimental_data_manifest.tsv"
    rows: list[tuple[object, ...]] = []
    for source_id in ("TNF", "IFNG", "TNF_IFNG", "T6_MATCHED", "WT", "KO1", "KO2", "TCR"):
        path = sources[source_id]
        rows.append(
            (
                source_id,
                "source_workbook",
                SOURCE_SPECS[source_id]["canonical_name"],
                sha256(path),
                path.stat().st_size,
                "local source; unchanged by preparation",
            )
        )
    for output in outputs:
        output_path = ROOT / str(output["path"])
        rows.append(
            (
                Path(str(output["path"])).stem,
                "canonical_analysis_table",
                str(output["path"]),
                output["sha256"],
                output["size_bytes"],
                "version-controlled author-generated processed data",
            )
        )
    write_tsv(
        manifest_path,
        (
            "record_id",
            "record_type",
            "filename_or_path",
            "sha256",
            "size_bytes",
            "repository_policy",
        ),
        rows,
    )
    return manifest_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-dir",
        type=Path,
        required=True,
        help="Directory containing the eight source workbooks",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Canonical output directory (default: {DEFAULT_OUTPUT})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_dir = args.input_dir.resolve()
    output_dir = args.output_dir.resolve()
    try:
        output_dir.relative_to(ROOT)
    except ValueError as error:
        raise ValueError("--output-dir must be located inside the repository root") from error
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    lock_path = output_dir.parent / f".{output_dir.name}.prepare.lock"
    lock_handle = acquire_lock(lock_path)
    try:
        remove_stale_temporary_files(output_dir)
        sources = {
            source_id: locate_one(input_dir, str(spec["pattern"]))
            for source_id, spec in SOURCE_SPECS.items()
        }

        cytokine_output, cytokine_qc_parts = prepare_cytokine_de(sources, output_dir)
        design_output, t6_outputs, t6_qc = prepare_t6_matched(
            sources["T6_MATCHED"], output_dir
        )
        singlecell_outputs, singlecell_qc = prepare_single_cell(sources, output_dir)
        outputs = [cytokine_output, design_output, *t6_outputs, *singlecell_outputs]
        manifest_path = write_manifest(sources, outputs, output_dir)

        for output in outputs:
            output_path = ROOT / str(output["path"])
            if (
                not output_path.is_file()
                or output_path.stat().st_size != output["size_bytes"]
                or sha256(output_path) != output["sha256"]
            ):
                raise RuntimeError(f"Post-write validation failed for {output_path}")

        qc = {
            "schema_version": 1,
            "cytokine_differential_expression": cytokine_qc_parts[0],
            "t6_matched_design": t6_qc,
            "single_cell": singlecell_qc,
            "canonical_outputs": outputs,
            "manifest_sha256": sha256(manifest_path),
        }
        qc_path = output_dir / "experimental_preparation_qc.json"
        write_text_atomic(
            qc_path, json.dumps(qc, indent=2, sort_keys=True) + "\n"
        )
    finally:
        fcntl.flock(lock_handle.fileno(), fcntl.LOCK_UN)
        lock_handle.close()

    print("EXPERIMENTAL_PREPARATION_OK")
    print(f"manifest={manifest_path}")
    print(f"qc={qc_path}")
    for output in outputs:
        print(
            f"{output['path']}\t{output['rows']} rows\t"
            f"sha256={output['sha256']}"
        )


if __name__ == "__main__":
    main()
