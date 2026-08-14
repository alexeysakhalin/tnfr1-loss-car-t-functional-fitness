#!/usr/bin/env python3
"""Prepare transparent cohort tables without a database import layer.

All sample-level outputs are written below data/analysis/, which is ignored by
Git. The optional CheckMate step writes only aggregate gene-model statistics.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import io
import json
import math
import os
import re
import sys
from datetime import date, datetime, time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable, Iterator, Sequence

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - dependency error is explicit
    raise SystemExit("openpyxl is required: python -m pip install openpyxl") from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RAW = ROOT / "data" / "raw"
DEFAULT_OUTPUT = ROOT / "data" / "analysis"
DEFAULT_MANIFEST = ROOT / "data" / "source_manifest.tsv"
DEFAULT_RESOURCES = ROOT / "resources"
BUFFER_SIZE = 1024 * 1024

EXPRESSION_COLUMNS = [
    "cohort_id",
    "sample_id",
    "gene_symbol",
    "expr_value",
    "expression_unit",
    "rank_percentile",
]
METADATA_COLUMNS = [
    "cohort_id",
    "sample_id",
    "trial_id",
    "treatment_arm",
    "os_time_months",
    "os_event",
    "pfs_time_months",
    "pfs_event",
    "tumor_purity",
]

FIXED_ANALYSIS_GENES = {
    "TNFRSF1A",
    "IFNGR1",
    "IFNG",
    "CD8A",
    "ICAM1",
    "IRF1",
    "B2M",
    "CD2",
    "CD3D",
    "CD3E",
    "CD8B",
    "TRAC",
    "STAT1",
    "CXCL9",
    "CXCL10",
    "NFKBIA",
    "TNFAIP3",
    "RELB",
    "BIRC3",
    "CXCL2",
    "CXCL3",
    "JUNB",
    "FOS",
    "EGR1",
    "TRAF1",
    "NFKB2",
    "IER3",
    "PTGS2",
    "IL6",
    "CCL2",
    "CCL20",
    "TNIP1",
}
# Five-gene bulk T-cell score shared by all four source matrices. TRAC is not
# present in three cohorts and is therefore not part of the cross-cohort score.
TCELL_GENES = ["CD2", "CD3D", "CD3E", "CD8A", "CD8B"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--raw-dir", type=Path, default=DEFAULT_RAW)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--resources-dir", type=Path, default=DEFAULT_RESOURCES)
    parser.add_argument(
        "--cohort",
        action="append",
        choices=["imvigor210", "su2c", "liu"],
        help="Open cohort to prepare; repeat as needed. Default: all three.",
    )
    parser.add_argument(
        "--include-checkmate-aggregates",
        action="store_true",
        help="Recalculate aggregate C6 gene models from the local Braun workbook.",
    )
    return parser.parse_args()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(BUFFER_SIZE), b""):
            digest.update(block)
    return digest.hexdigest()


def read_manifest(path: Path) -> dict[str, dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter="\t"))
    result = {row["source_id"]: row for row in rows}
    if len(result) != len(rows):
        raise ValueError("Duplicate source_id in source manifest")
    return result


def verified_input(
    raw_dir: Path, manifest: dict[str, dict[str, str]], source_id: str
) -> Path:
    if source_id not in manifest:
        raise KeyError(f"Source absent from manifest: {source_id}")
    record = manifest[source_id]
    path = raw_dir / record["expected_filename"]
    if not path.is_file():
        raise FileNotFoundError(f"Missing {source_id}: {path}")
    expected_size = int(record["size_bytes"])
    if path.stat().st_size != expected_size:
        raise RuntimeError(
            f"Size mismatch for {source_id}: {path.stat().st_size} != {expected_size}"
        )
    observed = sha256(path)
    if observed != record["sha256"]:
        raise RuntimeError(f"SHA-256 mismatch for {source_id}: {observed}")
    return path


def first_verified_input(
    raw_dir: Path,
    manifest: dict[str, dict[str, str]],
    source_ids: Sequence[str],
) -> tuple[str, Path]:
    for source_id in source_ids:
        record = manifest.get(source_id)
        if record and (raw_dir / record["expected_filename"]).is_file():
            return source_id, verified_input(raw_dir, manifest, source_id)
    expected = [manifest[source_id]["expected_filename"] for source_id in source_ids]
    raise FileNotFoundError(
        "None of the accepted source files is present: " + ", ".join(expected)
    )


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    return " ".join(text.split())


def as_float(value: object) -> float:
    if value is None:
        return math.nan
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value)
    if not text or text.upper() in {"NA", "N/A", "NAN", "NE"}:
        return math.nan
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return math.nan


def format_number(value: float) -> str:
    return "" if not math.isfinite(value) else format(value, ".15g")


def load_targets(resources_dir: Path) -> set[str]:
    targets = set(FIXED_ANALYSIS_GENES)
    for filename in ["CAR_T_state_signatures.csv", "Figure_5F_curated_gene_sets.csv"]:
        path = resources_dir / filename
        if not path.is_file():
            raise FileNotFoundError(path)
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                gene = clean(row.get("gene"))
                if gene:
                    targets.add(gene)
    return targets


def load_hgnc_mapping(
    path: Path, targets: set[str]
) -> tuple[dict[str, str], dict[str, str], dict[str, str], dict[str, object]]:
    """Load the frozen, analysis-specific subset of the HGNC snapshot."""
    with path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter="\t"))
    required = {
        "target_gene", "hgnc_symbol", "hgnc_id", "entrez_id",
        "ensembl_gene_id", "mapping_basis", "hgnc_snapshot_sha256",
        "hgnc_snapshot_retrieved",
    }
    if not rows or not required.issubset(rows[0]):
        raise RuntimeError("Frozen gene-identifier resource has an invalid schema")
    row_targets = {clean(row["target_gene"]) for row in rows}
    if row_targets != targets:
        raise RuntimeError(
            "Frozen identifier resource and current analysis gene sets differ; "
            "regenerate the resource from the declared HGNC snapshot."
        )

    source_symbol_to_target: dict[str, str] = {target: target for target in targets}
    entrez_to_target: dict[str, str] = {}
    ensembl_to_target: dict[str, str] = {}
    unresolved = []
    for row in rows:
        target = clean(row["target_gene"])
        current = clean(row.get("hgnc_symbol"))
        if current:
            source_symbol_to_target.setdefault(current, target)
        else:
            unresolved.append(target)
        entrez = clean(row.get("entrez_id"))
        ensembl = clean(row.get("ensembl_gene_id")).split(".")[0]
        if entrez:
            if entrez in entrez_to_target and entrez_to_target[entrez] != target:
                raise RuntimeError(f"One Entrez ID maps to multiple requested targets: {entrez}")
            entrez_to_target[entrez] = target
        if ensembl:
            if ensembl in ensembl_to_target and ensembl_to_target[ensembl] != target:
                raise RuntimeError(f"One Ensembl ID maps to multiple requested targets: {ensembl}")
            ensembl_to_target[ensembl] = target

    mapping_qc = {
        "targets_defined": len(targets),
        "targets_resolved_by_hgnc": len(targets) - len(unresolved),
        "targets_without_hgnc_record": sorted(unresolved),
        "mapping_rows": len(rows),
        "mapping_resource_sha256": sha256(path),
        "hgnc_snapshot_sha256": rows[0]["hgnc_snapshot_sha256"],
        "hgnc_snapshot_retrieved": rows[0]["hgnc_snapshot_retrieved"],
    }
    return source_symbol_to_target, entrez_to_target, ensembl_to_target, mapping_qc


def load_full_hgnc_mapping(path: Path) -> tuple[dict[str, str], dict[str, str], dict[str, object]]:
    with gzip.open(path, "rt", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter="\t"))
    entrez: dict[str, str] = {}
    ensembl: dict[str, str] = {}
    ambiguous_entrez: set[str] = set()
    ambiguous_ensembl: set[str] = set()
    for row in rows:
        symbol = clean(row.get("symbol"))
        entrez_id = clean(row.get("entrez_id"))
        ensembl_id = clean(row.get("ensembl_gene_id")).split(".")[0]
        if not symbol:
            continue
        if entrez_id:
            if entrez_id in entrez and entrez[entrez_id] != symbol:
                ambiguous_entrez.add(entrez_id)
            elif entrez_id not in ambiguous_entrez:
                entrez[entrez_id] = symbol
        if ensembl_id:
            if ensembl_id in ensembl and ensembl[ensembl_id] != symbol:
                ambiguous_ensembl.add(ensembl_id)
            elif ensembl_id not in ambiguous_ensembl:
                ensembl[ensembl_id] = symbol
    for identifier in ambiguous_entrez:
        entrez.pop(identifier, None)
    for identifier in ambiguous_ensembl:
        ensembl.pop(identifier, None)
    return entrez, ensembl, {
        "full_mapping_rows": len(rows),
        "full_mapping_resource_sha256": sha256(path),
        "ambiguous_entrez_ids_excluded": len(ambiguous_entrez),
        "ambiguous_ensembl_ids_excluded": len(ambiguous_ensembl),
    }


def atomic_gzip_dicts(path: Path, fieldnames: Sequence[str], rows: Iterable[dict[str, object]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".part")
    count = 0
    with temporary.open("wb") as binary:
        with gzip.GzipFile(filename="", mode="wb", fileobj=binary, mtime=0) as zipped:
            with io.TextIOWrapper(zipped, encoding="utf-8", newline="") as text:
                writer = csv.DictWriter(text, fieldnames=fieldnames, delimiter="\t", lineterminator="\n")
                writer.writeheader()
                for row in rows:
                    writer.writerow({name: row.get(name, "") for name in fieldnames})
                    count += 1
    os.replace(temporary, path)
    return count


def atomic_tsv(path: Path, fieldnames: Sequence[str], rows: Iterable[dict[str, object]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".part")
    count = 0
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})
            count += 1
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temporary, path)
    return count


def combine_gzip_tables(
    input_paths: Sequence[Path], output_path: Path, fieldnames: Sequence[str]
) -> int:
    def rows() -> Iterator[dict[str, str]]:
        for path in input_paths:
            with gzip.open(path, "rt", encoding="utf-8", newline="") as handle:
                reader = csv.DictReader(handle, delimiter="\t")
                if list(reader.fieldnames or []) != list(fieldnames):
                    raise RuntimeError(f"Unexpected intermediate schema: {path}")
                yield from reader

    return atomic_gzip_dicts(output_path, fieldnames, rows())


def expression_dicts(
    cohort_id: str,
    expression_unit: str,
    values: Iterable[tuple[str, str, float, float]],
) -> Iterator[dict[str, object]]:
    for sample_id, gene, value, rank_percentile in values:
        yield {
            "cohort_id": cohort_id,
            "sample_id": sample_id,
            "gene_symbol": gene,
            "expr_value": format_number(value),
            "expression_unit": expression_unit,
            "rank_percentile": format_number(rank_percentile),
        }


def transcriptome_rank_percentiles(matrix: object) -> object:
    """Average-tie percentile ranks within each sample over all source features."""
    import numpy as np
    from scipy import stats

    values = np.asarray(matrix, dtype=float)
    result = np.full(values.shape, np.nan, dtype=float)
    for sample_index in range(values.shape[1]):
        column = values[:, sample_index]
        finite = np.isfinite(column)
        n = int(finite.sum())
        if n == 0:
            continue
        if n == 1:
            result[finite, sample_index] = 0.5
            continue
        ranks = stats.rankdata(column[finite], method="average")
        result[finite, sample_index] = (ranks - 1.0) / (n - 1.0)
    return result


def selected_values_with_ranks(
    matrix: object,
    rank_matrix: object,
    source_targets: Sequence[str | None],
    sample_ids: Sequence[str],
) -> list[tuple[str, str, float, float]]:
    import numpy as np

    by_target: dict[str, list[int]] = defaultdict(list)
    for feature_index, target in enumerate(source_targets):
        if target:
            by_target[target].append(feature_index)
    rows: list[tuple[str, str, float, float]] = []
    for gene in sorted(by_target):
        indices = by_target[gene]
        expression = np.nanmean(matrix[indices, :], axis=0)
        ranks = np.nanmean(rank_matrix[indices, :], axis=0)
        for sample_index, sample_id in enumerate(sample_ids):
            if np.isfinite(expression[sample_index]) and np.isfinite(ranks[sample_index]):
                rows.append(
                    (sample_id, gene, float(expression[sample_index]), float(ranks[sample_index]))
                )
    return rows


def aggregate_transcriptome(
    matrix: object, source_gene_symbols: Sequence[str]
) -> tuple[list[str], object, dict[str, int]]:
    """Aggregate duplicate source features to one expression row per gene."""
    import numpy as np

    values = np.asarray(matrix, dtype=float)
    if values.shape[0] != len(source_gene_symbols):
        raise RuntimeError("Expression row count and source-gene mapping length differ")
    groups: dict[str, list[int]] = defaultdict(list)
    for index, symbol in enumerate(source_gene_symbols):
        if symbol:
            groups[symbol].append(index)
    genes = sorted(groups)
    aggregated = np.vstack(
        [np.nanmean(values[groups[gene], :], axis=0) for gene in genes]
    )
    counts = {gene: len(groups[gene]) for gene in genes}
    return genes, aggregated, counts


def prepare_imvigor(
    raw_dir: Path,
    output_dir: Path,
    manifest: dict[str, dict[str, str]],
    full_entrez_to_symbol: dict[str, str],
    source_symbol_to_target: dict[str, str],
) -> dict[str, object]:
    import numpy as np

    clinical_path = verified_input(raw_dir, manifest, "imvigor210_clinical_export")
    expression_path = verified_input(raw_dir, manifest, "imvigor210_expression_export")

    with clinical_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        raw_header = next(reader)
        header = ["sample_id", *raw_header[1:]]
        clinical = [dict(zip(header, row)) for row in reader if row and clean(row[0])]
    clinical_by_id = {clean(row["sample_id"]): row for row in clinical}
    if len(clinical_by_id) != 348:
        raise RuntimeError(f"IMvigor210 clinical QC failed: expected 348 samples, found {len(clinical_by_id)}")

    with expression_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        sample_ids = [clean(value) for value in header[1:]]
        if len(sample_ids) != 348 or len(set(sample_ids)) != 348:
            raise RuntimeError("IMvigor210 expression header must contain 348 unique samples")
        if set(sample_ids) != set(clinical_by_id):
            raise RuntimeError("IMvigor210 clinical/expression sample sets differ")
        source_gene_symbols: list[str] = []
        all_vectors: list[list[float]] = []
        for row in reader:
            if not row:
                continue
            source_gene_symbols.append(full_entrez_to_symbol.get(clean(row[0]), ""))
            all_vectors.append([as_float(value) for value in row[1:]])
    raw_matrix = np.asarray(all_vectors, dtype=float)
    mapped_genes, matrix, feature_counts = aggregate_transcriptome(
        raw_matrix, source_gene_symbols
    )
    source_targets = [source_symbol_to_target.get(gene) for gene in mapped_genes]
    ranks = transcriptome_rank_percentiles(matrix)
    expression_values = selected_values_with_ranks(matrix, ranks, source_targets, sample_ids)

    metadata = []
    for sample_id in sorted(sample_ids):
        metadata.append(
            {
                "cohort_id": "IMvigor210_BLCA",
                "sample_id": sample_id,
                "trial_id": "",
                "treatment_arm": "ATEZOLIZUMAB",
                "os_time_months": "",
                "os_event": "",
                "pfs_time_months": "",
                "pfs_event": "",
                "tumor_purity": "",
            }
        )

    expr_path = output_dir / "IMvigor210_BLCA.selected_expression.tsv.gz"
    meta_path = output_dir / "IMvigor210_BLCA.sample_metadata.tsv.gz"
    expression_rows = atomic_gzip_dicts(
        expr_path,
        EXPRESSION_COLUMNS,
        expression_dicts(
            "IMvigor210_BLCA",
            "log2CPM",
            expression_values,
        ),
    )
    metadata_rows = atomic_gzip_dicts(meta_path, METADATA_COLUMNS, metadata)
    return {
        "cohort_id": "IMvigor210_BLCA",
        "samples": len(sample_ids),
        "source_features_total": int(raw_matrix.shape[0]),
        "mapped_unique_genes_ranked": int(matrix.shape[0]),
        "unmapped_source_features": sum(not symbol for symbol in source_gene_symbols),
        "duplicate_source_features_aggregated": sum(count - 1 for count in feature_counts.values()),
        "selected_genes": len({row[1] for row in expression_values}),
        "expression_rows": expression_rows,
        "metadata_rows": metadata_rows,
        "expression_sha256": sha256(expr_path),
        "metadata_sha256": sha256(meta_path),
    }


def workbook_header(ws: object, row_number: int) -> list[str]:
    row = next(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True))
    return [clean(value) for value in row]


def prepare_su2c(
    raw_dir: Path,
    output_dir: Path,
    manifest: dict[str, dict[str, str]],
    targets: set[str],
    source_symbol_to_target: dict[str, str],
    full_ensembl_to_symbol: dict[str, str],
) -> dict[str, object]:
    import numpy as np

    workbook_path = verified_input(raw_dir, manifest, "su2c_mark_supplement")
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    required_sheets = {"Table_S1_Clinical_Annotations", "Table_S13_RNA_TPM"}
    if not required_sheets.issubset(wb.sheetnames):
        raise RuntimeError("SU2C workbook lacks required S1/S13 sheets")

    clinical_ws = wb["Table_S1_Clinical_Annotations"]
    clinical_header = workbook_header(clinical_ws, 3)
    clinical_index = {name: index for index, name in enumerate(clinical_header) if name}
    required_columns = {
        "Harmonized_SU2C_RNA_Tumor_Sample_ID_v2",
        "Pre-treatment_RNA_Sample_QC",
        "Agent_PD1",
        "Harmonized_Confirmed_BOR",
    }
    if not required_columns.issubset(clinical_index):
        raise RuntimeError(f"SU2C S1 lacks columns: {sorted(required_columns.difference(clinical_index))}")
    clinical_by_id: dict[str, dict[str, object]] = {}
    for values in clinical_ws.iter_rows(min_row=4, values_only=True):
        sample_id = clean(values[clinical_index["Harmonized_SU2C_RNA_Tumor_Sample_ID_v2"]])
        if not sample_id:
            continue
        if sample_id in clinical_by_id:
            raise RuntimeError(f"Duplicate SU2C RNA sample ID: {sample_id}")
        clinical_by_id[sample_id] = {
            name: values[index] if index < len(values) else None for name, index in clinical_index.items()
        }

    expression_ws = wb["Table_S13_RNA_TPM"]
    expression_header = workbook_header(expression_ws, 3)
    sample_ids = expression_header[2:]
    if len(sample_ids) != 152 or len(set(sample_ids)) != 152:
        raise RuntimeError("SU2C S13 must contain 152 unique RNA samples")
    if set(sample_ids) != set(clinical_by_id):
        raise RuntimeError("SU2C S1/S13 RNA sample sets differ")

    source_gene_symbols: list[str] = []
    all_vectors: list[list[float]] = []
    remapped_from_ensembl: set[str] = set()
    spreadsheet_formatted_symbol_rows = 0
    for row in expression_ws.iter_rows(min_row=4, values_only=True):
        ensembl = clean(row[0]).split(".")[0]
        displayed_value = row[1]
        if isinstance(displayed_value, (date, datetime, time)):
            displayed_symbol = ""
            spreadsheet_formatted_symbol_rows += 1
        else:
            displayed_symbol = clean(displayed_value)
        mapped_symbol = full_ensembl_to_symbol.get(ensembl) or displayed_symbol
        source_gene_symbols.append(mapped_symbol)
        target = source_symbol_to_target.get(mapped_symbol)
        if target and mapped_symbol != displayed_symbol:
            remapped_from_ensembl.add(target)
        all_vectors.append([as_float(value) for value in row[2 : 2 + len(sample_ids)]])
    wb.close()
    raw_matrix = np.asarray(all_vectors, dtype=float)
    mapped_genes, matrix, feature_counts = aggregate_transcriptome(
        raw_matrix, source_gene_symbols
    )
    source_targets = [source_symbol_to_target.get(gene) for gene in mapped_genes]
    ranks = transcriptome_rank_percentiles(matrix)
    expression_values = selected_values_with_ranks(matrix, ranks, source_targets, sample_ids)

    qc_counts = Counter(clean(row["Pre-treatment_RNA_Sample_QC"]) for row in clinical_by_id.values())
    if qc_counts != Counter({"Keep": 136, "Flag": 16}):
        raise RuntimeError(f"Unexpected SU2C RNA QC counts: {dict(qc_counts)}")
    metadata = []
    for sample_id in sorted(sample_ids):
        row = clinical_by_id[sample_id]
        metadata.append(
            {
                "cohort_id": "SU2C_MARK_NSCLC",
                "sample_id": sample_id,
                "trial_id": "",
                "treatment_arm": (clean(row.get("Agent_PD1")) or "anti-PD-(L)1").upper(),
                "os_time_months": "",
                "os_event": "",
                "pfs_time_months": "",
                "pfs_event": "",
                "tumor_purity": "",
            }
        )

    expr_path = output_dir / "SU2C_MARK_NSCLC.selected_expression.tsv.gz"
    meta_path = output_dir / "SU2C_MARK_NSCLC.sample_metadata.tsv.gz"
    expression_rows = atomic_gzip_dicts(
        expr_path,
        EXPRESSION_COLUMNS,
        expression_dicts(
            "SU2C_MARK_NSCLC",
            "TPM",
            expression_values,
        ),
    )
    metadata_rows = atomic_gzip_dicts(meta_path, METADATA_COLUMNS, metadata)
    return {
        "cohort_id": "SU2C_MARK_NSCLC",
        "samples": len(sample_ids),
        "source_features_total": int(raw_matrix.shape[0]),
        "mapped_unique_genes_ranked": int(matrix.shape[0]),
        "unmapped_source_features": sum(not symbol for symbol in source_gene_symbols),
        "duplicate_source_features_aggregated": sum(count - 1 for count in feature_counts.values()),
        "selected_genes": len({row[1] for row in expression_values}),
        "expression_rows": expression_rows,
        "metadata_rows": metadata_rows,
        "rna_qc_counts": dict(qc_counts),
        "targets_recovered_by_identifier_mapping": sorted(remapped_from_ensembl),
        "spreadsheet_formatted_symbol_rows": spreadsheet_formatted_symbol_rows,
        "expression_sha256": sha256(expr_path),
        "metadata_sha256": sha256(meta_path),
    }


def prepare_liu(
    raw_dir: Path,
    output_dir: Path,
    manifest: dict[str, dict[str, str]],
    targets: set[str],
    source_symbol_to_target: dict[str, str],
) -> dict[str, object]:
    import numpy as np

    clinical_path = verified_input(raw_dir, manifest, "liu2019_clinical_supplement")
    expression_source_id, expression_path = first_verified_input(
        raw_dir,
        manifest,
        ["liu2019_tpm_supplement", "liu2019_expression_export"],
    )
    wb = openpyxl.load_workbook(clinical_path, read_only=True, data_only=True)
    sheet = "Supplemental Table 1"
    if sheet not in wb.sheetnames:
        raise RuntimeError("Liu workbook lacks Supplemental Table 1")
    ws = wb[sheet]
    header_values = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    header = ["sample_id" if index == 0 else clean(value) for index, value in enumerate(header_values)]
    clinical_by_id: dict[str, dict[str, object]] = {}
    rejected_nonpatient_rows = 0
    for values in ws.iter_rows(min_row=4, values_only=True):
        sample_id = clean(values[0] if values else None)
        if not re.fullmatch(r"Patient[0-9]+", sample_id):
            if any(value is not None for value in values):
                rejected_nonpatient_rows += 1
            continue
        if sample_id in clinical_by_id:
            raise RuntimeError(f"Duplicate Liu patient ID: {sample_id}")
        clinical_by_id[sample_id] = dict(zip(header, values))
    wb.close()
    if len(clinical_by_id) != 144:
        raise RuntimeError(f"Liu clinical QC failed: expected 144 patients, found {len(clinical_by_id)}")

    delimiter = "\t" if expression_path.suffix.lower() == ".txt" else ","
    with expression_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        raw_header = next(reader)
        source_genes = [clean(value) for value in raw_header[1:]]
        sample_ids: list[str] = []
        sample_vectors: list[list[float]] = []
        for values in reader:
            if not values:
                continue
            sample_id = clean(values[0])
            if sample_id in sample_ids:
                raise RuntimeError(f"Duplicate Liu expression sample: {sample_id}")
            sample_ids.append(sample_id)
            sample_vectors.append([as_float(value) for value in values[1:]])
    if len(sample_ids) != 121 or len(set(sample_ids)) != 121:
        raise RuntimeError("Liu TPM supplement must contain 121 unique samples")
    if not set(sample_ids).issubset(clinical_by_id):
        raise RuntimeError("One or more Liu TPM samples are absent from the clinical table")

    sample_matrix = np.asarray(sample_vectors, dtype=float)
    raw_feature_matrix = sample_matrix.T
    mapped_genes, feature_matrix, feature_counts = aggregate_transcriptome(
        raw_feature_matrix, source_genes
    )
    source_targets = [source_symbol_to_target.get(gene) for gene in mapped_genes]
    rank_matrix = transcriptome_rank_percentiles(feature_matrix)
    expression_values = selected_values_with_ranks(
        feature_matrix, rank_matrix, source_targets, sample_ids
    )
    all_genes = sorted({row[1] for row in expression_values})

    metadata = []
    for sample_id in sorted(sample_ids):
        row = clinical_by_id[sample_id]
        metadata.append(
            {
                "cohort_id": "LIU2019_MELANOMA",
                "sample_id": sample_id,
                "trial_id": "",
                "treatment_arm": (clean(row.get("Tx")) or "anti-PD-1").upper(),
                "os_time_months": "",
                "os_event": "",
                "pfs_time_months": "",
                "pfs_event": "",
                "tumor_purity": format_number(as_float(row.get("purity"))),
            }
        )

    expr_path = output_dir / "LIU2019_MELANOMA.selected_expression.tsv.gz"
    meta_path = output_dir / "LIU2019_MELANOMA.sample_metadata.tsv.gz"
    expression_rows = atomic_gzip_dicts(
        expr_path,
        EXPRESSION_COLUMNS,
        expression_dicts(
            "LIU2019_MELANOMA", "TPM", expression_values
        ),
    )
    metadata_rows = atomic_gzip_dicts(meta_path, METADATA_COLUMNS, metadata)
    return {
        "cohort_id": "LIU2019_MELANOMA",
        "clinical_patient_rows": len(clinical_by_id),
        "rejected_nonpatient_rows": rejected_nonpatient_rows,
        "samples": len(sample_ids),
        "clinical_without_tpm": len(set(clinical_by_id).difference(sample_ids)),
        "selected_genes": len(all_genes),
        "source_features_total": int(raw_feature_matrix.shape[0]),
        "mapped_unique_genes_ranked": int(feature_matrix.shape[0]),
        "duplicate_source_features_aggregated": sum(count - 1 for count in feature_counts.values()),
        "expression_source_id": expression_source_id,
        "expression_rows": expression_rows,
        "metadata_rows": metadata_rows,
        "expression_sha256": sha256(expr_path),
        "metadata_sha256": sha256(meta_path),
    }


def marker_set(resources_dir: Path, cluster: int) -> list[str]:
    path = resources_dir / "CAR_T_state_signatures.csv"
    selected: list[tuple[float, int, str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for order, row in enumerate(csv.DictReader(handle)):
            if clean(row.get("cluster")).removeprefix("C") != str(cluster):
                continue
            gene = clean(row.get("gene"))
            if gene:
                selected.append((as_float(row.get("avg_log2FC")), order, gene))
    selected.sort(key=lambda item: (-item[0] if math.isfinite(item[0]) else math.inf, item[1]))
    genes: list[str] = []
    for _, _, gene in selected:
        if gene not in genes:
            genes.append(gene)
        if len(genes) == 20:
            break
    if len(genes) < 3:
        raise RuntimeError(f"Insufficient genes for C{cluster}")
    return genes


def safe_z_numpy(values: object) -> object:
    import numpy as np

    array = np.asarray(values, dtype=float)
    mean = np.nanmean(array)
    sd = np.nanstd(array, ddof=1)
    if not np.isfinite(sd) or sd == 0:
        return np.zeros(array.shape, dtype=float)
    return (array - mean) / sd


def bh_adjust(p_values: list[float]) -> list[float]:
    import numpy as np

    values = np.asarray(p_values, dtype=float)
    result = np.full(values.shape, np.nan)
    ok = np.isfinite(values)
    finite = values[ok]
    order = np.argsort(finite)
    ranked = finite[order]
    adjusted = ranked * len(ranked) / np.arange(1, len(ranked) + 1)
    adjusted = np.minimum.accumulate(adjusted[::-1])[::-1]
    restored = np.empty_like(adjusted)
    restored[order] = np.minimum(adjusted, 1.0)
    result[np.where(ok)[0]] = restored
    return result.tolist()


def resolve_braun_sources(
    raw_dir: Path, manifest: dict[str, dict[str, str]]
) -> tuple[Path, str, Path, str, list[str]]:
    combined_record = manifest["braun_checkmate_supplement"]
    combined_path = raw_dir / combined_record["expected_filename"]
    if combined_path.is_file():
        verified = verified_input(raw_dir, manifest, "braun_checkmate_supplement")
        return (
            verified,
            "S1_Clinical_and_Immune_Data",
            verified,
            "S4A_RNA_Expression",
            ["braun_checkmate_supplement"],
        )
    clinical = verified_input(raw_dir, manifest, "braun_checkmate_clinical_split")
    expression = verified_input(raw_dir, manifest, "braun_checkmate_expression_split")
    return (
        clinical,
        "Clinical_and_Immune_Data",
        expression,
        "RNA_Expression",
        ["braun_checkmate_clinical_split", "braun_checkmate_expression_split"],
    )


def read_braun_clinical(
    workbook_path: Path, sheet: str
) -> tuple[list[dict[str, object]], list[str]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise RuntimeError("Braun workbook lacks S1 clinical sheet")
    ws = wb[sheet]
    header = workbook_header(ws, 2)
    index = {name: pos for pos, name in enumerate(header) if name}
    required = {"RNA_ID", "Cohort", "Arm", "OS_CNSR", "PFS_CNSR"}
    if not required.issubset(index):
        raise RuntimeError(f"Braun S1 lacks columns: {sorted(required.difference(index))}")
    rows = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        sample_id = clean(values[index["RNA_ID"]])
        if not sample_id or sample_id == "NA":
            continue
        rows.append({name: values[pos] if pos < len(values) else None for name, pos in index.items()})
    wb.close()
    if len(rows) != 311 or len({clean(row["RNA_ID"]) for row in rows}) != 311:
        raise RuntimeError("Braun source must contain 311 unique RNA-profiled tumors")
    arm_counts = Counter(clean(row["Arm"]).upper() for row in rows)
    if arm_counts != Counter({"NIVOLUMAB": 181, "EVEROLIMUS": 130}):
        raise RuntimeError(f"Unexpected CheckMate arm counts: {dict(arm_counts)}")
    all_os_events = sum(as_float(row.get("OS_CNSR")) == 1 for row in rows)
    all_pfs_events = sum(as_float(row.get("PFS_CNSR")) == 1 for row in rows)
    if (all_os_events, all_pfs_events) != (231, 276):
        raise RuntimeError(
            f"Unexpected all-arm event counts: OS={all_os_events}, PFS={all_pfs_events}"
        )
    nivo = [row for row in rows if clean(row["Arm"]).upper() == "NIVOLUMAB"]
    if len(nivo) != 181:
        raise RuntimeError(f"Expected 181 nivolumab RNA tumors; found {len(nivo)}")
    trial_counts = Counter(clean(row["Cohort"]) for row in nivo)
    if trial_counts != Counter({"CM-009": 16, "CM-010": 45, "CM-025": 120}):
        raise RuntimeError(f"Unexpected nivolumab trial counts: {dict(trial_counts)}")
    os_events = sum(as_float(row.get("OS_CNSR")) == 1 for row in nivo)
    pfs_events = sum(as_float(row.get("PFS_CNSR")) == 1 for row in nivo)
    if (os_events, pfs_events) != (123, 159):
        raise RuntimeError(f"Unexpected nivolumab event counts: OS={os_events}, PFS={pfs_events}")
    return rows, [clean(row["RNA_ID"]) for row in rows]


def braun_expression_layout(
    workbook_path: Path, sheet: str, sample_ids: list[str]
) -> tuple[object, object, list[int]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise RuntimeError("Braun workbook lacks S4A expression sheet")
    ws = wb[sheet]
    header = workbook_header(ws, 2)
    positions = {sample: index for index, sample in enumerate(header) if index > 0 and sample}
    missing = sorted(set(sample_ids).difference(positions))
    if missing:
        wb.close()
        raise RuntimeError(f"Braun expression matrix lacks RNA samples: {missing[:5]}")
    return wb, ws, [positions[sample] for sample in sample_ids]


def prepare_checkmate_aggregates(
    raw_dir: Path,
    output_dir: Path,
    manifest: dict[str, dict[str, str]],
    resources_dir: Path,
    targets: set[str],
    source_symbol_to_target: dict[str, str],
) -> dict[str, object]:
    import numpy as np
    from scipy import stats

    (
        clinical_path,
        clinical_sheet,
        expression_path,
        expression_sheet,
        source_ids,
    ) = resolve_braun_sources(raw_dir, manifest)
    clinical, sample_ids = read_braun_clinical(clinical_path, clinical_sheet)
    trials_all = np.array([clean(row["Cohort"]) for row in clinical])
    arms_all = np.array([clean(row["Arm"]).upper() for row in clinical])
    nivolumab_indices = np.where(arms_all == "NIVOLUMAB")[0]
    trials = trials_all[nivolumab_indices]
    c6_genes = marker_set(resources_dir, 6)
    score_genes = set(c6_genes).union(TCELL_GENES)

    wb, ws, selected_positions = braun_expression_layout(
        expression_path, expression_sheet, sample_ids
    )
    raw_source_genes: list[str] = []
    raw_vectors: list[list[float]] = []
    spreadsheet_formatted_gene_rows = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        raw_gene = row[0]
        if isinstance(raw_gene, (date, datetime, time)):
            gene = ""
            spreadsheet_formatted_gene_rows += 1
        else:
            gene = clean(raw_gene)
        raw_source_genes.append(gene)
        raw_vectors.append([as_float(row[position]) for position in selected_positions])
    wb.close()

    raw_matrix = np.asarray(raw_vectors, dtype=float)
    source_genes, matrix, feature_counts = aggregate_transcriptome(
        raw_matrix, raw_source_genes
    )
    target_by_source_gene = [source_symbol_to_target.get(gene) for gene in source_genes]
    rank_matrix = transcriptome_rank_percentiles(matrix)
    selected_expression = selected_values_with_ranks(
        matrix, rank_matrix, target_by_source_gene, sample_ids
    )
    processed_dir = output_dir.parent / "processed"
    processed_dir.mkdir(parents=True, exist_ok=True)
    checkmate_expression_path = processed_dir / "checkmate_selected_expression.tsv.gz"
    checkmate_expression_rows = atomic_gzip_dicts(
        checkmate_expression_path,
        EXPRESSION_COLUMNS,
        expression_dicts(
            "CHECKMATE_CCRCC", "normalized_expression", selected_expression
        ),
    )
    metadata_rows = []
    for row in clinical:
        metadata_rows.append(
            {
                "cohort_id": "CHECKMATE_CCRCC",
                "sample_id": clean(row["RNA_ID"]),
                "trial_id": clean(row["Cohort"]),
                "treatment_arm": clean(row["Arm"]).upper(),
                "os_time_months": format_number(as_float(row.get("OS"))),
                "os_event": format_number(as_float(row.get("OS_CNSR"))),
                "pfs_time_months": format_number(as_float(row.get("PFS"))),
                "pfs_event": format_number(as_float(row.get("PFS_CNSR"))),
                "tumor_purity": "",
            }
        )
    checkmate_metadata_path = processed_dir / "checkmate_sample_metadata.tsv.gz"
    checkmate_metadata_rows = atomic_gzip_dicts(
        checkmate_metadata_path, METADATA_COLUMNS, metadata_rows
    )

    gene_to_index = {gene: index for index, gene in enumerate(source_genes)}
    combined_scores = {
        gene: matrix[gene_to_index[gene], nivolumab_indices]
        for gene in score_genes
        if gene in gene_to_index
    }
    c6_present = [gene for gene in c6_genes if gene in combined_scores]
    tcell_present = [gene for gene in TCELL_GENES if gene in combined_scores]
    if len(c6_present) < 3 or len(tcell_present) < 4:
        raise RuntimeError("Insufficient C6 or T-cell gene coverage in Braun expression")
    c6_matrix = np.column_stack([safe_z_numpy(combined_scores[gene]) for gene in c6_present])
    tcell_matrix = np.column_stack([safe_z_numpy(combined_scores[gene]) for gene in tcell_present])
    c6_score = np.nanmean(c6_matrix, axis=1)
    tcell_score_z = safe_z_numpy(np.nanmean(tcell_matrix, axis=1))

    trial_levels = sorted(set(trials))
    trial_dummies = np.column_stack([(trials == level).astype(float) for level in trial_levels[1:]])
    residual_design = np.column_stack([np.ones(len(trials)), tcell_score_z, trial_dummies])
    residual_beta = np.linalg.lstsq(residual_design, c6_score, rcond=None)[0]
    c6_residual = c6_score - residual_design @ residual_beta
    median = float(np.median(c6_residual))
    group = (c6_residual > median).astype(int)
    if sorted(Counter(group).values()) != [90, 91]:
        raise RuntimeError(f"Unexpected residualized C6 group sizes: {dict(Counter(group))}")

    design = np.column_stack([np.ones(len(trials)), group, tcell_score_z, trial_dummies])
    excluded = set(c6_present).union(tcell_present)
    models: list[dict[str, object]] = []
    seen_genes: set[str] = set()
    for feature_index, gene in enumerate(source_genes):
        if gene in excluded or gene in seen_genes:
            continue
        seen_genes.add(gene)
        y = matrix[feature_index, nivolumab_indices]
        complete = np.isfinite(y) & np.all(np.isfinite(design), axis=1)
        n = int(complete.sum())
        if n < 163 or len(set(group[complete])) != 2:
            continue
        x = design[complete]
        outcome = y[complete]
        rank = np.linalg.matrix_rank(x)
        if rank != x.shape[1]:
            continue
        beta = np.linalg.lstsq(x, outcome, rcond=None)[0]
        residuals = outcome - x @ beta
        degrees_freedom = n - x.shape[1]
        variance = float(residuals @ residuals / degrees_freedom)
        covariance = variance * np.linalg.inv(x.T @ x)
        standard_error = math.sqrt(float(covariance[1, 1]))
        if not math.isfinite(standard_error) or standard_error == 0:
            continue
        statistic = float(beta[1] / standard_error)
        p_value = float(2 * stats.t.sf(abs(statistic), degrees_freedom))
        critical = float(stats.t.ppf(0.975, degrees_freedom))
        high = outcome[group[complete] == 1]
        low = outcome[group[complete] == 0]
        models.append(
            {
                "gene": gene,
                "n": n,
                "beta": float(beta[1]),
                "se": standard_error,
                "ci_low": float(beta[1] - critical * standard_error),
                "ci_high": float(beta[1] + critical * standard_error),
                "p": p_value,
                "mean_high": float(np.mean(high)),
                "mean_low": float(np.mean(low)),
            }
        )
    if len(models) < 1000:
        raise RuntimeError(f"Unexpectedly small CheckMate model family: {len(models)}")
    adjusted = bh_adjust([float(row["p"]) for row in models])
    for row, value in zip(models, adjusted):
        row["BH_p"] = value
        row["neglog10_p"] = -math.log10(max(float(row["p"]), sys.float_info.min))
        row["direction"] = (
            "higher in adjusted C6-high" if float(row["beta"]) >= 0 else "higher in adjusted C6-low"
        )
    models.sort(key=lambda row: (float(row["p"]), str(row["gene"])))
    model_path = output_dir / "checkmate_c6_global_gene_models.tsv.gz"
    model_columns = [
        "gene",
        "n",
        "beta",
        "se",
        "ci_low",
        "ci_high",
        "p",
        "mean_high",
        "mean_low",
        "BH_p",
        "neglog10_p",
        "direction",
    ]
    model_rows = atomic_gzip_dicts(
        model_path,
        model_columns,
        (
            {
                key: format_number(float(row[key])) if key in {
                    "beta", "se", "ci_low", "ci_high", "p", "mean_high", "mean_low", "BH_p", "neglog10_p"
                } else row[key]
                for key in model_columns
            }
            for row in models
        ),
    )

    low_values = tcell_score_z[group == 0]
    high_values = tcell_score_z[group == 1]
    welch_p = float(stats.ttest_ind(low_values, high_values, equal_var=False).pvalue)
    balance_rows = []
    for label, values in [("Low", low_values), ("High", high_values)]:
        balance_rows.append(
            {
                "c6_group": label,
                "n": len(values),
                "tcell_score_mean": format_number(float(np.mean(values))),
                "tcell_score_sd": format_number(float(np.std(values, ddof=1))),
                "welch_p": format_number(welch_p),
            }
        )
    balance_path = output_dir / "checkmate_c6_group_balance.tsv"
    atomic_tsv(
        balance_path,
        ["c6_group", "n", "tcell_score_mean", "tcell_score_sd", "welch_p"],
        balance_rows,
    )

    qc = {
        "source_ids": source_ids,
        "source_sha256": {
            source_id: sha256(
                clinical_path if "clinical" in source_id else expression_path
            )
            for source_id in source_ids
        },
        "analysis_population": "nivolumab-treated RNA-profiled tumors",
        "n_all_rna": len(sample_ids),
        "n_primary_nivolumab": len(nivolumab_indices),
        "trial_counts": dict(sorted(Counter(trials).items())),
        "group_counts": {
            "C6-low adjusted": int((group == 0).sum()),
            "C6-high adjusted": int((group == 1).sum()),
        },
        "c6_genes_defined": c6_genes,
        "c6_genes_present": c6_present,
        "tcell_genes_defined": TCELL_GENES,
        "tcell_genes_present": tcell_present,
        "excluded_outcome_genes": sorted(excluded),
        "global_model_family_size": len(models),
        "multiplicity": "Benjamini-Hochberg across all fitted outcome genes",
        "model": "expression ~ C6_group + TcellScore_z + factor(trial)",
        "source_features_total": int(raw_matrix.shape[0]),
        "mapped_unique_genes_ranked": int(matrix.shape[0]),
        "duplicate_source_features_aggregated": sum(
            count - 1 for count in feature_counts.values()
        ),
        "spreadsheet_formatted_gene_rows_skipped": spreadsheet_formatted_gene_rows,
        "checkmate_selected_expression_rows": checkmate_expression_rows,
        "checkmate_sample_metadata_rows": checkmate_metadata_rows,
        "checkmate_selected_expression_sha256": sha256(checkmate_expression_path),
        "checkmate_sample_metadata_sha256": sha256(checkmate_metadata_path),
        "model_table_sha256": sha256(model_path),
        "balance_table_sha256": sha256(balance_path),
    }
    qc_path = output_dir / "checkmate_c6_aggregate_qc.json"
    qc_path.write_text(json.dumps(qc, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return {
        "checkmate_selected_expression_rows": checkmate_expression_rows,
        "checkmate_sample_metadata_rows": checkmate_metadata_rows,
        "checkmate_selected_expression_sha256": sha256(checkmate_expression_path),
        "checkmate_sample_metadata_sha256": sha256(checkmate_metadata_path),
        "checkmate_aggregate_models": model_rows,
        "checkmate_model_table_sha256": sha256(model_path),
        "checkmate_balance_table_sha256": sha256(balance_path),
        "checkmate_qc_sha256": sha256(qc_path),
    }


def main() -> int:
    args = parse_args()
    try:
        raw_dir = args.raw_dir.resolve()
        output_dir = args.output_dir.resolve()
        resources_dir = args.resources_dir.resolve()
        manifest = read_manifest(args.manifest.resolve())
        targets = load_targets(resources_dir)
        selected_mapping_path = resources_dir / "analysis_gene_identifiers.tsv"
        full_mapping_path = resources_dir / "hgnc_20260814_gene_identifiers.tsv.gz"
        source_symbols, _, _, mapping_qc = load_hgnc_mapping(
            selected_mapping_path, targets
        )
        full_entrez, full_ensembl, full_mapping_qc = load_full_hgnc_mapping(
            full_mapping_path
        )
        mapping_qc.update(full_mapping_qc)
        selected_cohorts = args.cohort or ["imvigor210", "su2c", "liu"]
        output_dir.mkdir(parents=True, exist_ok=True)
        cohort_qc = []
        for cohort in selected_cohorts:
            if cohort == "imvigor210":
                cohort_qc.append(
                    prepare_imvigor(
                        raw_dir, output_dir, manifest, full_entrez, source_symbols
                    )
                )
            elif cohort == "su2c":
                cohort_qc.append(
                    prepare_su2c(
                        raw_dir,
                        output_dir,
                        manifest,
                        targets,
                        source_symbols,
                        full_ensembl,
                    )
                )
            elif cohort == "liu":
                cohort_qc.append(
                    prepare_liu(raw_dir, output_dir, manifest, targets, source_symbols)
                )

        cohort_ids = {
            "imvigor210": "IMvigor210_BLCA",
            "su2c": "SU2C_MARK_NSCLC",
            "liu": "LIU2019_MELANOMA",
        }
        expression_parts = [
            output_dir / f"{cohort_ids[cohort]}.selected_expression.tsv.gz"
            for cohort in selected_cohorts
        ]
        metadata_parts = [
            output_dir / f"{cohort_ids[cohort]}.sample_metadata.tsv.gz"
            for cohort in selected_cohorts
        ]
        combined_expression = output_dir / "open_cohort_selected_expression.tsv.gz"
        combined_metadata = output_dir / "open_cohort_sample_metadata.tsv.gz"
        combined_expression_rows = combine_gzip_tables(
            expression_parts, combined_expression, EXPRESSION_COLUMNS
        )
        combined_metadata_rows = combine_gzip_tables(
            metadata_parts, combined_metadata, METADATA_COLUMNS
        )
        for intermediate in [*expression_parts, *metadata_parts]:
            intermediate.unlink()

        qc: dict[str, object] = {
            "schema_version": 2,
            "targets_defined": len(targets),
            "target_gene_sha256": hashlib.sha256(
                ("\n".join(sorted(targets)) + "\n").encode("utf-8")
            ).hexdigest(),
            "identifier_mapping": mapping_qc,
            "cohorts": cohort_qc,
            "combined_open_expression_rows": combined_expression_rows,
            "combined_open_metadata_rows": combined_metadata_rows,
            "combined_open_expression_sha256": sha256(combined_expression),
            "combined_open_metadata_sha256": sha256(combined_metadata),
        }
        if args.include_checkmate_aggregates:
            qc.update(
                prepare_checkmate_aggregates(
                    raw_dir,
                    output_dir,
                    manifest,
                    resources_dir,
                    targets,
                    source_symbols,
                )
            )
        qc_path = output_dir / "open_cohort_preparation_qc.json"
        qc_path.write_text(json.dumps(qc, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        print(f"Prepared {len(cohort_qc)} open cohort(s); QC: {qc_path}")
        return 0
    except (OSError, ValueError, KeyError, RuntimeError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
